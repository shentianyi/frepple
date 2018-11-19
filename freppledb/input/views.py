#
# Copyright (C) 2007-2013 by frePPLe bvba
#
# This library is free software; you can redistribute it and/or modify it
# under the terms of the GNU Affero General Public License as published
# by the Free Software Foundation; either version 3 of the License, or
# (at your option) any later version.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Affero
# General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
import csv
import math
import sys
import traceback
from datetime import datetime
from io import BytesIO, StringIO

from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.contenttypes.models import ContentType
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.db import connections, transaction
from django.db.models import Q
from django.db.models.fields import CharField
from django.http import HttpResponse, Http404
from django.http.response import StreamingHttpResponse, HttpResponseServerError, HttpResponseBadRequest, JsonResponse
from django.shortcuts import render
from django.template import loader
from django.utils.decorators import method_decorator
from django.utils.formats import get_format
from django.utils.translation import ugettext_lazy as _
from django.utils.translation import ungettext
from django.utils.translation import string_concat
from django.utils.encoding import force_text
from django.utils.text import format_lazy
from django.views.generic import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import NamedStyle, PatternFill

from freppledb.boot import getAttributeFields
from freppledb.common.message.responsemessage import ResponseMessage
from freppledb.common.models import Parameter, Comment, Bucket
from freppledb.common.utils import la_enum, la_field
from freppledb.common.utils.la_field import decimal2float
from freppledb.input.forms import ForecastUploadForm
from freppledb.input.models import Resource, Operation, Location, SetupMatrix, SetupRule, ItemSuccessor, ItemCustomer, \
    ForecastYear, ForecastVersion, Forecast, ForecastCommentOperation
from freppledb.input.models import Skill, Buffer, Customer, Demand, DeliveryOrder
from freppledb.input.models import Item, OperationResource, OperationMaterial
from freppledb.input.models import Calendar, CalendarBucket, ManufacturingOrder, SubOperation
from freppledb.input.models import ResourceSkill, Supplier, ItemSupplier, searchmode
from freppledb.input.models import ItemDistribution, DistributionOrder, PurchaseOrder
from freppledb.input.models import OperationPlan, OperationPlanMaterial, OperationPlanResource
from freppledb.common.report import GridReport, GridFieldBool, GridFieldLastModified, GridFieldCreateOrUpdateDate, \
    GridFieldDate
from freppledb.common.report import GridFieldDateTime, GridFieldTime, GridFieldText
from freppledb.common.report import GridFieldNumber, GridFieldInteger, GridFieldCurrency
from freppledb.common.report import GridFieldChoice, GridFieldDuration
from freppledb.admin import data_site
from django.utils import timezone
import json
from django.core.serializers.json import DjangoJSONEncoder

import logging

from freppledb.input.uploader import ForecastUploader

logger = logging.getLogger(__name__)


@staff_member_required
def search(request):
    term = request.GET.get('term').strip()
    result = []

    # Loop over all models in the data_site
    # We are interested in models satisfying these criteria:
    #  - primary key is of type text
    #  - user has change permissions
    for cls, admn in data_site._registry.items():
        if request.user.has_perm("%s.view_%s" % (cls._meta.app_label, cls._meta.object_name.lower())) and isinstance(
                cls._meta.pk, CharField):
            query = cls.objects.using(request.database).filter(pk__icontains=term).order_by('pk').values_list('pk')
            count = len(query)
            if count > 0:
                result.append({'value': None, 'label': (ungettext(
                    '%(name)s - %(count)d match',
                    '%(name)s - %(count)d matches', count) % {'name': force_text(cls._meta.verbose_name),
                                                              'count': count}).capitalize()
                               })
                result.extend([{
                    'url': "/detail/%s/%s/" % (cls._meta.app_label, cls._meta.object_name.lower()),
                    'value': i[0]
                } for i in query[:10]])

    # Construct reply
    return HttpResponse(
        content_type='application/json; charset=%s' % settings.DEFAULT_CHARSET,
        content=json.dumps(result).encode(settings.DEFAULT_CHARSET)
    )


class PathReport(GridReport):
    '''
    A report showing the upstream supply path or following downstream a
    where-used path.
    The supply path report shows all the materials, operations and resources
    used to make a certain item.
    The where-used report shows all the materials and operations that use
    a specific item.
    '''
    template = 'input/path.html'
    title = ''
    filterable = False
    frozenColumns = 0
    editable = False
    default_sort = None
    multiselect = False
    help_url = 'user-guide/user-interface/supply-path-where-used.html'

    rows = (
        GridFieldText('depth', title=_('depth'), editable=False, sortable=False),
        GridFieldText('operation', title=_('operation'), editable=False, sortable=False, formatter='detail',
                      extra='"role":"input/operation"'),
        GridFieldNumber('quantity', title=_('quantity'), editable=False, sortable=False),
        GridFieldText('location', title=_('location'), editable=False, sortable=False),
        GridFieldText('type', title=_('type'), editable=False, sortable=False),
        GridFieldDuration('duration', title=_('duration'), editable=False, sortable=False),
        GridFieldDuration('duration_per', title=_('duration per unit'), editable=False, sortable=False),
        GridFieldText('resources', editable=False, sortable=False, extra='formatter:reslistfmt'),
        GridFieldText('buffers', editable=False, sortable=False, hidden=True),
        GridFieldText('suboperation', editable=False, sortable=False, hidden=True),
        GridFieldText('numsuboperations', editable=False, sortable=False, hidden=True),
        GridFieldText('parentoper', editable=False, sortable=False, hidden=True),
        GridFieldText('realdepth', editable=False, sortable=False, hidden=True),
        GridFieldText('id', editable=False, sortable=False, hidden=True),
        GridFieldText('parent', editable=False, sortable=False, hidden=True),
        GridFieldText('leaf', editable=False, sortable=False, hidden=True),
        GridFieldText('expanded', editable=False, sortable=False, hidden=True),
    )

    # Attributes to be specified by the subclasses
    objecttype = None
    downstream = None

    @classmethod
    def basequeryset(reportclass, request, *args, **kwargs):
        return reportclass.objecttype.objects.filter(name__exact=args[0]).values('name')

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        if reportclass.downstream:
            request.session['lasttab'] = 'whereused'
        else:
            request.session['lasttab'] = 'supplypath'
        return {
            'title': force_text(reportclass.objecttype._meta.verbose_name) + " " + args[0],
            'post_title': _("where used") if reportclass.downstream else _("supply path"),
            'downstream': reportclass.downstream,
            'active_tab': reportclass.downstream and 'whereused' or 'supplypath',
            'model': reportclass.objecttype._meta
        }

    @classmethod
    def getRoot(reportclass, request, entity):
        raise Http404("invalid entity type")

    @classmethod
    def findDeliveries(reportclass, item, location, db):
        # Automatically detect delivery operations. This is done by looking for
        # a buffer for this item and location combination.
        buf = None
        # Find a buffer record
        for b in Buffer.objects.using(db).filter(item=item, location=location):
            buf = b
        if not buf:
            # Create a buffer record
            buf = Buffer(
                name='%s @ %s' % (item, location),
                item=Item.objects.using(db).get(name=item),
                location=Location.objects.using(db).get(name=location)
            )
        return reportclass.findReplenishment(buf, db, 0, 1, 0, False)

    @classmethod
    def findUsage(reportclass, buffer, db, level, curqty, realdepth, pushsuper):
        result = [
            (level - 1, None, i.operation, curqty, 0, None, realdepth, pushsuper,
             buffer.location.nr if buffer.location else None)
            for i in buffer.item.operationmaterials.filter(quantity__lt=0).filter(
                operation__location__name=buffer.location.nr).only('operation').using(db)
        ]
        for i in ItemDistribution.objects.using(db).filter(
                item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft,
                origin__name=buffer.location.nr
        ):
            i.item = buffer.item
            result.append(
                (level - 1, None, i, curqty, 0, None, realdepth - 1, pushsuper, i.location.nr if i.location else None))
        return result

    @classmethod
    def findReplenishment(reportclass, buffer, db, level, curqty, realdepth, pushsuper):
        # If a producing operation is set on the buffer, we use that and skip the
        # automated search described below.
        # If no producing operation is set, we look for item distribution and
        # item supplier models for the item and location combination. (As a special
        # case in case only a single location exists in the model, a match on the
        # item is sufficient).
        result = []
        if Location.objects.using(db).count() > 1:
            # Multiple locations
            for i in ItemSupplier.objects.using(db).filter(
                    Q(location__isnull=True) | Q(location__name=buffer.location.nr),
                    item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft
            ):
                i.item = buffer.item
                i.location = buffer.location
                result.append(
                    (level, None, i, curqty, 0, None, realdepth, pushsuper,
                     buffer.location.nr if buffer.location else None)
                )
            for i in ItemDistribution.objects.using(db).filter(
                    Q(location__isnull=True) | Q(location__name=buffer.location.nr),
                    item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft
            ):
                i.item = buffer.item
                i.location = buffer.location
                result.append(
                    (level, None, i, curqty, 0, None, realdepth, pushsuper, i.location.nr if i.location else None)
                )
            for i in Operation.objects.using(db).filter(
                    Q(location__isnull=True) | Q(location__name=buffer.location.nr),
                    item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft
            ):
                i.item = buffer.item
                i.location = buffer.location
                result.append(
                    (level, None, i, curqty, 0, None, realdepth, pushsuper, i.location.nr if i.location else None)
                )
        else:
            # Single location
            for i in ItemSupplier.objects.using(db).filter(
                    item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft
            ):
                i.item = buffer.item
                i.location = buffer.location
                result.append(
                    (level, None, i, curqty, 0, None, realdepth, pushsuper,
                     buffer.location.nr if buffer.location else None)
                )
            for i in Operation.objects.using(db).filter(
                    item__lft__lte=buffer.item.lft, item__rght__gt=buffer.item.lft
            ):
                i.item = buffer.item
                i.location = buffer.location
                result.append(
                    (level, None, i, curqty, 0, None, realdepth, pushsuper,
                     buffer.location.nr if buffer.location else None)
                )
        return result

    @classmethod
    def query(reportclass, request, basequery):
        '''
        A function that recurses upstream or downstream in the supply chain.
        '''
        # Update item and location hierarchies
        Item.rebuildHierarchy(database=request.database)
        Location.rebuildHierarchy(database=request.database)

        entity = basequery.query.get_compiler(basequery.db).as_sql(with_col_aliases=False)[1]
        entity = entity[0]
        root = reportclass.getRoot(request, entity)

        # Recurse over all operations
        # TODO the current logic isn't generic enough. A lot of buffers may not be explicitly
        # defined, and are created on the fly by deliveries, itemsuppliers or itemdistributions.
        # Currently we don't account for such situations.
        # TODO usage search doesn't find item distributions from that location
        counter = 1
        operations = set()
        while len(root) > 0:
            # Pop the current node from the stack
            level, parent, curoperation, curqty, issuboperation, parentoper, realdepth, pushsuper, location = root.pop()
            curnode = counter
            counter += 1
            if isinstance(location, str):
                curlocation = Location.objects.all().using(request.database).get(name=location)

            # If an operation has parent operations we forget about the current operation
            # and use only the parent
            if pushsuper and not isinstance(curoperation, (ItemSupplier, ItemDistribution)):
                hasParents = False
                for x in curoperation.superoperations.using(request.database).only('operation').order_by("-priority"):
                    root.append(
                        (level, parent, x.operation, curqty, issuboperation, parentoper, realdepth, False, location))
                    hasParents = True
                if hasParents:
                    continue

            # Avoid showing the same operation twice.
            # This feature is enabled by default. Without it we cannot correctly display
            # supply paths with loops (which are normally a modeling error).
            # The use of this feature has some drawbacks  a) because it is not intuitive
            # to understand where operations are skipped in the path, and b) because
            # the quantity of each occurrence might be different.
            # You may choose can disable this feature by commenting out the next 3 lines.
            if curoperation in operations:
                continue
            operations.add(curoperation)

            # Find the next level
            hasChildren = False
            subcount = 0
            if reportclass.downstream:
                # Downstream recursion
                if isinstance(curoperation, ItemSupplier):
                    name = 'Purchase %s @ %s from %s' % (curoperation.item.name, location, curoperation.supplier.name)
                    optype = "purchase"
                    duration = curoperation.leadtime
                    duration_per = None
                    buffers = [("%s @ %s" % (curoperation.item.name, curoperation.location.nr), 1), ]
                    if curoperation.resource:
                        resources = [(curoperation.resource.name, float(curoperation.resource_qty))]
                    else:
                        resources = None
                    try:
                        downstr = Buffer.objects.using(request.database).get(
                            name="%s @ %s" % (curoperation.item.name, curoperation.location.nr))
                        root.extend(
                            reportclass.findUsage(downstr, request.database, level, curqty, realdepth + 1, True))
                    except Buffer.DoesNotExist:
                        downstr = Buffer(name="%s @ %s" % (curoperation.item.name, curoperation.location.nr),
                                         item=curoperation.item, location=curlocation)
                        root.extend(
                            reportclass.findUsage(downstr, request.database, level, curqty, realdepth + 1, True))
                elif isinstance(curoperation, ItemDistribution):
                    name = 'Ship %s from %s to %s' % (
                        curoperation.item.name, curoperation.origin.name, curoperation.location.nr)
                    optype = "distribution"
                    duration = curoperation.leadtime
                    duration_per = None
                    buffers = [
                        ("%s @ %s" % (curoperation.item.name, curoperation.origin.name), -1),
                        ("%s @ %s" % (curoperation.item.name, curoperation.location.nr), 1)
                    ]
                    if curoperation.resource:
                        resources = [(curoperation.resource.name, float(curoperation.resource_qty))]
                    else:
                        resources = None
                    try:
                        downstr = Buffer.objects.using(request.database).get(
                            name="%s @ %s" % (curoperation.item.name, location))
                        root.extend(
                            reportclass.findUsage(downstr, request.database, level, curqty, realdepth + 1, True))
                    except Buffer.DoesNotExist:
                        downstr = Buffer(name="%s @ %s" % (curoperation.item.name, location), item=curoperation.item,
                                         location=curlocation)
                        root.extend(
                            reportclass.findUsage(downstr, request.database, level, curqty, realdepth + 1, True))
                else:
                    name = curoperation.name
                    optype = curoperation.type
                    duration = curoperation.duration
                    duration_per = curoperation.duration_per
                    buffers = [('%s @ %s' % (x.item.name, curoperation.location.nr), float(x.quantity)) for x in
                               curoperation.operationmaterials.only('item', 'quantity').using(request.database)]
                    resources = [(x.resource.name, float(x.quantity)) for x in
                                 curoperation.operationresources.only('resource', 'quantity').using(request.database)]
                    for x in curoperation.operationmaterials.filter(quantity__gt=0).only('item').using(
                            request.database):
                        curflows = x.item.operationmaterials.filter(quantity__lt=0,
                                                                    operation__location=curoperation.location.nr).only(
                            'operation', 'quantity').using(request.database)
                        for y in curflows:
                            hasChildren = True
                            root.append((level - 1, curnode, y.operation, - curqty * y.quantity, subcount, None,
                                         realdepth - 1, pushsuper,
                                         x.operation.location.nr if x.operation.location else None))
                        try:
                            downstr = Buffer.objects.using(request.database).get(
                                name="%s @ %s" % (x.item.name, location))
                            root.extend(
                                reportclass.findUsage(downstr, request.database, level - 1, curqty, realdepth - 1,
                                                      True))
                        except Buffer.DoesNotExist:
                            downstr = Buffer(name="%s @ %s" % (curoperation.item.name, location), item=x.item,
                                             location=curlocation)
                            root.extend(
                                reportclass.findUsage(downstr, request.database, level - 1, curqty, realdepth - 1,
                                                      True))
                    for x in curoperation.suboperations.using(request.database).only('suboperation').order_by(
                            "-priority"):
                        subcount += curoperation.type == "routing" and 1 or -1
                        root.append((level - 1, curnode, x.suboperation, curqty, subcount, curoperation, realdepth,
                                     False, location))
                        hasChildren = True
            else:
                # Upstream recursion
                if isinstance(curoperation, ItemSupplier):
                    name = 'Purchase %s @ %s from %s' % (curoperation.item.name, location, curoperation.supplier.name)
                    optype = "purchase"
                    duration = curoperation.leadtime
                    duration_per = None
                    buffers = [("%s @ %s" % (curoperation.item.name, location), 1), ]
                    if curoperation.resource:
                        resources = [(curoperation.resource.name, float(curoperation.resource_qty))]
                    else:
                        resources = None
                elif isinstance(curoperation, ItemDistribution):
                    name = 'Ship %s from %s to %s' % (curoperation.item.name, curoperation.origin.name, location)
                    optype = "distribution"
                    duration = curoperation.leadtime
                    duration_per = None
                    buffers = [
                        ("%s @ %s" % (curoperation.item.name, curoperation.origin.name), -1),
                        ("%s @ %s" % (curoperation.item.name, curoperation.location.nr), 1)
                    ]
                    if curoperation.resource:
                        resources = [(curoperation.resource.name, float(curoperation.resource_qty))]
                    else:
                        resources = None
                    try:
                        upstr = Buffer.objects.using(request.database).get(
                            name="%s @ %s" % (curoperation.item.name, curoperation.origin.name))
                        root.extend(
                            reportclass.findReplenishment(upstr, request.database, level + 2, curqty, realdepth + 1,
                                                          True))
                    except Buffer.DoesNotExist:
                        upstr = Buffer(name="%s @ %s" % (curoperation.item.name, curoperation.origin.name),
                                       item=curoperation.item, location=curoperation.origin)
                        root.extend(
                            reportclass.findReplenishment(upstr, request.database, level + 2, curqty, realdepth + 1,
                                                          True))
                else:
                    name = curoperation.name
                    optype = curoperation.type
                    duration = curoperation.duration
                    duration_per = curoperation.duration_per
                    buffers = [('%s @ %s' % (x.item.name, curoperation.location.nr), float(x.quantity)) for x in
                               curoperation.operationmaterials.only('item', 'quantity').using(request.database)]
                    resources = [(x.resource.name, float(x.quantity)) for x in
                                 curoperation.operationresources.only('resource', 'quantity').using(request.database)]
                    curflows = curoperation.operationmaterials.filter(quantity__lt=0).only('item', 'quantity').using(
                        request.database)
                    for y in curflows:
                        b = Buffer(
                            name='%s @ %s' % (y.item.name, curoperation.location.nr),
                            item=y.item,
                            location=curoperation.location
                        )
                        root.extend(
                            reportclass.findReplenishment(b, request.database, level + 2, curqty, realdepth + 1, True))
                    for x in curoperation.suboperations.using(request.database).only('suboperation').order_by(
                            "-priority"):
                        subcount += curoperation.type == "routing" and 1 or -1
                        root.append((level + 1, curnode, x.suboperation, curqty, subcount, curoperation, realdepth,
                                     False, location))
                        hasChildren = True

            # Process the current node
            yield {
                'depth': abs(level),
                'id': curnode,
                'operation': name,
                'type': optype,
                'location': curoperation.location and curoperation.location.nr or '',
                'duration': duration,
                'duration_per': duration_per,
                'quantity': curqty,
                'suboperation': issuboperation,
                'buffers': buffers,
                'resources': resources,
                'parentoper': parentoper and parentoper.name,
                'parent': parent,
                'leaf': hasChildren and 'false' or 'true',
                'expanded': 'true',
                'numsuboperations': subcount,
                'realdepth': realdepth
            }


class UpstreamDemandPath(PathReport):
    downstream = False
    objecttype = Demand

    @classmethod
    def getRoot(reportclass, request, entity):
        from django.core.exceptions import ObjectDoesNotExist

        try:
            dmd = Demand.objects.using(request.database).get(name=entity)
        except ObjectDoesNotExist:
            raise Http404("demand %s doesn't exist" % entity)

        if dmd.operation:
            # Delivery operation on the demand
            return [(0, None, dmd.operation, 1, 0, None, 0, False, None)]
        else:
            # Autogenerated delivery operation
            try:
                return reportclass.findDeliveries(dmd.item, dmd.location, request.database)
            except:
                raise Http404("No supply path defined for demand %s" % entity)


class UpstreamItemPath(PathReport):
    downstream = False
    objecttype = Item

    @classmethod
    def getRoot(reportclass, request, entity):
        from django.core.exceptions import ObjectDoesNotExist
        try:
            locs = set()
            result = []
            it = Item.objects.using(request.database).get(name=entity)
            if reportclass.downstream:
                # Find all buffers where the item is being stored and walk downstream
                for b in Buffer.objects.filter(item=it).using(request.database):
                    locs.add(b.location.nr)
                    result.extend(reportclass.findUsage(b, request.database, 0, 1, 0, True))
                # Add item locations that can be replenished
                for itmdist in ItemDistribution.objects.using(request.database).filter(
                        item__lft__lte=it.lft, item__rght__gt=it.lft
                ):
                    if itmdist.location.nr in locs:
                        continue
                    locs.add(itmdist.location.nr)
                    itmdist.item = it
                    result.append(
                        (0, None, itmdist, 1, 0, None, 0, False, itmdist.location.nr)
                    )
                # Add item locations that can be replenished
                for itmsup in Operation.objects.using(request.database).filter(
                        item__lft__lte=it.lft, item__rght__gt=it.lft
                ):
                    if itmsup.location.nr in locs:
                        continue
                    locs.add(itmsup.location.nr)
                    itmsup.item = it
                    result.append(
                        (0, None, itmsup, 1, 0, None, 0, False, itmsup.location.nr)
                    )
                return result
            else:
                # Find the supply path of all buffers of this item
                for b in Buffer.objects.filter(item=entity).using(request.database):
                    result.extend(reportclass.findReplenishment(b, request.database, 0, 1, 0, True))
                # Add item locations that can be replenished
                for itmdist in ItemDistribution.objects.using(request.database).filter(
                        item__lft__lte=it.lft, item__rght__gt=it.lft
                ):
                    if itmdist.location.nr in locs:
                        continue
                    locs.add(itmdist.location.nr)
                    itmdist.item = it
                    result.append(
                        (0, None, itmdist, 1, 0, None, 0, False, itmdist.location.nr)
                    )
                # Add item locations that can be replenished
                for itmsup in Operation.objects.using(request.database).filter(
                        item__lft__lte=it.lft, item__rght__gt=it.lft
                ):
                    if itmsup.location.nr in locs:
                        continue
                    locs.add(itmsup.location.nr)
                    itmsup.item = it
                    result.append(
                        (0, None, itmsup, 1, 0, None, 0, False, itmsup.location.nr)
                    )
                return result
        except ObjectDoesNotExist:
            raise Http404("item %s doesn't exist" % entity)


class UpstreamBufferPath(PathReport):
    downstream = False
    objecttype = Buffer

    @classmethod
    def getRoot(reportclass, request, entity):
        from django.core.exceptions import ObjectDoesNotExist
        try:
            buf = Buffer.objects.using(request.database).get(name=entity)
            if reportclass.downstream:
                return reportclass.findUsage(buf, request.database, 0, 1, 0, True)
            else:
                return reportclass.findReplenishment(buf, request.database, 0, 1, 0, True)
        except ObjectDoesNotExist:
            raise Http404("buffer %s doesn't exist" % entity)


class UpstreamResourcePath(PathReport):
    downstream = False
    objecttype = Resource

    @classmethod
    def getRoot(reportclass, request, entity):
        from django.core.exceptions import ObjectDoesNotExist
        try:
            root = Resource.objects.using(request.database).get(name=entity)
        except ObjectDoesNotExist:
            raise Http404("resource %s doesn't exist" % entity)
        return [
            (0, None, i.operation, 1, 0, None, 0, True, i.operation.location.nr if i.operation.location else None)
            for i in root.operationresources.using(request.database).all()
        ]


class UpstreamOperationPath(PathReport):
    downstream = False
    objecttype = Operation

    @classmethod
    def getRoot(reportclass, request, entity):
        from django.core.exceptions import ObjectDoesNotExist
        try:
            oper = Operation.objects.using(request.database).get(name=entity)
            return [(0, None, oper, 1, 0, None, 0, True, oper.location.nr if oper.location else None)]
        except ObjectDoesNotExist:
            raise Http404("operation %s doesn't exist" % entity)


class DownstreamItemPath(UpstreamItemPath):
    downstream = True
    objecttype = Item


class DownstreamDemandPath(UpstreamDemandPath):
    downstream = True
    objecttype = Demand


class DownstreamBufferPath(UpstreamBufferPath):
    downstream = True
    objecttype = Buffer


class DownstreamResourcePath(UpstreamResourcePath):
    downstream = True
    objecttype = Resource


class DownstreamOperationPath(UpstreamOperationPath):
    downstream = True
    objecttype = Operation


# TODO　重点设计
class BufferList(GridReport):
    '''
    A list report to show buffers.
    '''
    title = _("buffers")
    basequeryset = Buffer.objects.all()
    model = Buffer
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/master-data/buffers.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('name', title=_('name'), key=True, formatter='detail',
                      extra='"role":"input/buffer", "editable":false'),
        GridFieldText('description', title=_('description')),
        GridFieldText('category', title=_('category'), initially_hidden=True),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True),
        GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
                      extra='"role":"input/item"'),
        GridFieldNumber('onhand', title=_('onhand')),
        GridFieldChoice('type', title=_('type'), choices=Buffer.types),
        GridFieldNumber('minimum', title=_('minimum')),
        GridFieldText('minimum_calendar', title=_('minimum calendar'), field_name='minimum_calendar__name',
                      formatter='detail', extra='"role":"input/calendar"', initially_hidden=True),
        GridFieldText('source', title=_('source')),
        GridFieldLastModified('lastmodified'),
        # Optional fields referencing the item
        GridFieldText(
            'item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the location
        GridFieldText(
            'location__description', title=string_concat(_('location'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__category', title=string_concat(_('location'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__available', title=string_concat(_('location'), ' - ', _('available')),
            initially_hidden=True, field_name='location__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
            initially_hidden=True, field_name='location__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'location__source', title=string_concat(_('location'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
    )


class SetupMatrixList(GridReport):
    '''
    A list report to show setup matrices.
    '''
    title = _("setup matrices")
    basequeryset = SetupMatrix.objects.all()
    model = SetupMatrix
    frozenColumns = 1
    help_url = 'user-guide/model-reference/setup-matrices.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/setupmatrix"',
                      editable=False),
        GridFieldText('nr', title=_('nr'), editable=False),
        GridFieldText('name', title=_('name'), key=True, formatter='detail', editable=False),
        GridFieldText('category', title=_('category'), editable=False),
        GridFieldText('subcategory', title=_('subcategory'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),

    )


class SetupRuleList(GridReport):
    '''
    A list report to show setup matrix rules.
    '''
    title = _("setup rules")
    basequeryset = SetupRule.objects.all()
    model = SetupRule
    frozenColumns = 1
    help_url = 'user-guide/model-reference/setup-matrices.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/setuprule"',
                         editable=False),
        GridFieldText('setupmatrix', title=_('setup matrix'), formatter='detail', extra='"role":"input/setupmatrix"',
                      editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        GridFieldText('fromsetup', title=_('from setup'), editable=False),
        GridFieldText('tosetup', title=_('to setup'), editable=False),
        GridFieldCurrency('cost', title=_('cost'), initially_hidden=True, editable=False),
        GridFieldDuration('duration', title=_('duration'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='id', editable=False, hidden=True),

    )


class ResourceList(GridReport):
    '''
    A list report to show resources.
    '''
    title = _("resources")
    basequeryset = Resource.objects.all()
    model = Resource
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/manufacturing-capacity/resources.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/resource"',
                      editable=False),
        GridFieldText('nr', title=_('nr'), editable=False),
        GridFieldText('name', title=_('name'), formatter='detail', editable=False),
        GridFieldText('owner_display', title=_('owner_display'), field_name='owner__nr', editable=False,
                      formatter='detail', extra='"role":"input/resource"'),
        GridFieldText('owner', title=_('owner_id'), field_name='owner_id', editable=False, hidden=True),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldChoice('type', title=_('type'), choices=Resource.types, editable=False),
        GridFieldNumber('maximum', title=_('maximum'), editable=False),

        GridFieldText('maximum_calendar', title=_('maximum calendar'), field_name='maximum_calendar__name',
                      formatter='detail', extra='"role":"input/calendar"', editable=False),

        GridFieldText('available', title=_('available'), field_name='available__name', formatter='detail',
                      extra='"role":"input/calendar"', editable=False),

        GridFieldText('location_display', title=_('location_display'), field_name='location__nr', editable=False),
        GridFieldInteger('location', title=_('location_id'), field_name='location_id', editable=False, hidden=True),

        GridFieldCurrency('cost', title=_('cost'), editable=False),

        GridFieldNumber('efficiency', title=_('efficiency %'), formatter='currency',
                        extra='"formatoptions":{"suffix":" %","defaultValue":"100.00"}', editable=False),
        GridFieldText(
            'efficiency_calendar', title=_('efficiency % calendar'), field_name='efficiency_calendar__name',
            formatter='detail', extra='"role":"input/calendar"', editable=False
        ),

        GridFieldText('setupmatrix_display', title=_('setupmatrix_display'), field_name='setupmatrix__nr',
                      editable=False),
        GridFieldText('setupmatrix', title=_('setupmatrix_id'), field_name='setupmatrix_id', editable=False,
                      hidden=True),
        GridFieldText('now_setup', title=_('now setup'), editable=False),

        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),

        # GridFieldText('owner', title=_('owner'), field_name='owner__name', formatter='detail',
        #               extra='"role":"input/resource"', initially_hidden=True, editable=False),
        # GridFieldDuration('maxearly', title=_('maxearly'), initially_hidden=True, editable=False),
        # GridFieldText('setupmatrix', title=_('setup matrix'), field_name='setupmatrix__name', formatter='detail',
        #               extra='"role":"input/setupmatrix"', initially_hidden=True, editable=False),
        # GridFieldText('source', title=_('source'), editable=False),
        # # Optional fields referencing the location
        # GridFieldText(
        #     'location__description', title=string_concat(_('location'), ' - ', _('description')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__category', title=string_concat(_('location'), ' - ', _('category')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__available', title=string_concat(_('location'), ' - ', _('available')),
        #     initially_hidden=True, field_name='location__available__name', formatter='detail',
        #     extra='"role":"input/calendar"', editable=False
        # ),
        # GridFieldText(
        #     'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
        #     initially_hidden=True, field_name='location__owner__name', formatter='detail',
        #     extra='"role":"input/location"', editable=False
        # ),
        # GridFieldText(
        #     'location__source', title=string_concat(_('location'), ' - ', _('source')),
        #     initially_hidden=True, editable=False
        # ),
    )


class LocationList(GridReport):
    '''
    A list report to show locations.
    '''
    title = _("locations")
    basequeryset = Location.objects.all()
    model = Location
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/master-data/locations.html'

    rows = (
        # . Translators: Translation included with Django
        # TODO CMARK id 的key作为连接到详情页面, 如何在nr这个属性上做链接,并且传递的不是nr的值, 而是id的值
        # key 表示链接的关键字
        # formatter 表示链接的动作
        # extra 表示其它的内容

        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/location"',
                      editable=False),
        GridFieldText('nr', title=_('nr'), editable=False),
        GridFieldText('name', title=_('name'), editable=False),
        # GridFieldText('name', title=_('name'), key=True, formatter='detail', extra='"role":"input/location"'),

        GridFieldText('area', title=_('area'), editable=False),
        GridFieldText('source', title=_('source'), editable=False),
        GridFieldText('available', title=_('available'), field_name='available__name', editable=False),

        # 新建一个显示列
        GridFieldText('owner_display', title=_('owner_display'), field_name='owner__nr', editable=False),

        # 因为是id 让外键永远不显示
        GridFieldText('owner', title=_('owner_id'), field_name='owner_id', editable=False, hidden=True),

        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        # GridFieldLastModified('lastmodified'),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),
    )


class CustomerList(GridReport):
    '''
    A list report to show customers.
    '''
    title = _("customers")
    basequeryset = Customer.objects.all()
    model = Customer
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/master-data/customers.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/customer"',
                      editable=False),
        GridFieldText('nr', title=_('customer nr'), editable=False),
        GridFieldText('name', title=_('name'), editable=False),
        GridFieldText('area', title=_('area'), editable=False),
        GridFieldText('address', title=_('address'), editable=False),
        GridFieldText('ship_address', title=_('ship address'), editable=False),
        GridFieldText('source', title=_('source'), editable=False),
        GridFieldText('available', title=_('available'), field_name='available__name', editable=False),
        GridFieldText('owner_display', title=_('owner_display'), field_name='owner__nr', editable=False),
        GridFieldText('owner', title=_('owner_id'), field_name='owner_id', editable=False, hidden=True),
        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        # GridFieldLastModified('lastmodified', title=_('lastmodified'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # CMARK 必须有为了弹框查询
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),
    )


class SupplierList(GridReport):
    '''
    A list report to show supplier.
    '''
    title = _("suppliers")
    basequeryset = Supplier.objects.all()
    model = Supplier
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/purchasing/suppliers.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/supplier"',
                      editable=False),
        GridFieldText('nr', title=_('supplier nr'), editable=False),
        GridFieldText('name', title=_('name'), editable=False),
        GridFieldText('area', title=_('area'), editable=False),
        GridFieldText('address', title=_('address'), editable=False),
        GridFieldText('ship_address', title=_('ship address'), editable=False),
        GridFieldText('country', title=_('country'), editable=False),
        GridFieldText('city', title=_('city'), editable=False),
        GridFieldText('phone', title=_('phone'), editable=False),
        GridFieldText('telephone', title=_('telephone'), editable=False),
        GridFieldText('contact', title=_('contact'), editable=False),
        GridFieldText('email', title=_('email'), editable=False),
        GridFieldText('source', title=_('source'), editable=False),
        GridFieldText('available', title=_('available'), field_name='available__name', editable=False),
        GridFieldText('owner_display', title=_('owner_display'), field_name='owner__nr', editable=False),
        GridFieldText('owner', title=_('owner_id'), field_name='owner_id', editable=False, hidden=True),
        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # CMARK 必须有为了弹框查询
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),
    )


class ItemSupplierList(GridReport):
    '''
    A list report to show item suppliers.
    '''
    title = _("item suppliers")
    basequeryset = ItemSupplier.objects.all()
    model = ItemSupplier
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/purchasing/item-suppliers.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/itemsupplier"',
                         editable=False),
        # 新建一个显示列
        GridFieldText('item_display', title=_('item_display'), field_name='item__nr', editable=False),
        GridFieldText('supplier_display', title=_('supplier_display'), field_name='supplier__nr', editable=False),
        GridFieldText('location_display', title=_('location_display'), field_name='location__nr', editable=False),

        # 因为是id 让外键永远不显示
        GridFieldInteger('item', title=_('item'), field_name='item__name', formatter='detail', editable=False,
                         hidden=True),
        GridFieldInteger('supplier', title=_('supplier'), field_name='supplier__name', formatter='detail',
                         editable=False, hidden=True),
        GridFieldInteger('location', title=_('location'), field_name='location__name', formatter='detail',
                         editable=False, hidden=True),
        GridFieldText('supplier_item_nr', title=_('supplier item nr'), editable=False),
        GridFieldText('status', title=_('status'), editable=False),
        GridFieldCurrency('cost', title=_('cost'), editable=False),
        GridFieldText('monetary_unit', title=_('monetary unit'), editable=False),
        GridFieldNumber('cost_unit', title=_('cost unit'), editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        GridFieldNumber('ratio', title=_('ratio'),
                        extra='"formatoptions":{"suffix":" %","defaultValue":"100.00"}', editable=False),

        GridFieldNumber('moq', title=_('MOQ'), editable=False),
        GridFieldNumber('order_unit_qty', title=_('order unit qty'), editable=False),
        GridFieldNumber('order_max_qty', title=_('order max qty'), editable=False),
        GridFieldNumber('product_time', title=_('product time'), editable=False),
        GridFieldNumber('load_time', title=_('load time'), editable=False),
        GridFieldNumber('transit_time', title=_('transit time'), editable=False),
        GridFieldNumber('receive_time', title=_('receive time'), editable=False),
        GridFieldInteger('mpq', title=_('mpq'), editable=False),
        GridFieldDate('earliest_order_date', title=_('earliest order date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_supplier_date', title=_('plan supplier date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_load_date', title=_('plan load date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_receive_date', title=_('plan receive date'), editable=False, initially_hidden=True),
        GridFieldInteger('outer_package_num', title=_('outer package num'), editable=False, initially_hidden=True),
        GridFieldInteger('pallet_num', title=_('pallet num'), editable=False, initially_hidden=True),
        GridFieldNumber('outer_package_gross_weight', title=_('outer package gross weight'), editable=False,
                        initially_hidden=True),
        GridFieldNumber('pallet_gross_weight', title=_('pallet gross weight'), editable=False, initially_hidden=True),
        GridFieldNumber('outer_package_volume', title=_('outer package volume'), editable=False, initially_hidden=True),
        GridFieldNumber('pallet_volume', title=_('pallet volume'), editable=False, initially_hidden=True),
        GridFieldDate('plan_list_date', title=_('plan list date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_delist_date', title=_('plan delist date'), editable=False, initially_hidden=True),
        GridFieldText('origin_country', title=_('origin country'), editable=False),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail', extra='"role":"input/location"'),
        # GridFieldDuration('leadtime', title=_('lead time')),
        # GridFieldNumber('sizeminimum', title=_('size minimum')),
        # GridFieldNumber('sizemultiple', title=_('size multiple')),
        # GridFieldCurrency('cost', title=_('cost')),
        # GridFieldDuration('fence', title=_('fence'), initially_hidden=True),
        # GridFieldText('resource', title=_('resource'), field_name='resource__name', formatter='detail', extra='"role":"input/resource"', initially_hidden=True),
        # GridFieldNumber('resource_qty', title=_('resource quantity'), initially_hidden=True),
        # GridFieldText('source', title=_('source')),
        # # Optional fields referencing the item
        # GridFieldText(
        #   'item__description', title=string_concat(_('item'), ' - ', _('description')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'item__category', title=string_concat(_('item'), ' - ', _('category')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
        #   field_name='item__owner__name', initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'item__source', title=string_concat(_('item'), ' - ', _('source')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldLastModified(
        #   'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
        #   initially_hidden=True, editable=False
        #   ),
        # # Optional fields referencing the location
        # GridFieldText(
        #   'location__description', title=string_concat(_('location'), ' - ', _('description')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'location__category', title=string_concat(_('location'), ' - ', _('category')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldText(
        #   'location__available', title=string_concat(_('location'), ' - ', _('available')),
        #   initially_hidden=True, field_name='location__available__name', formatter='detail',
        #   extra='"role":"input/calendar"', editable=False
        #   ),
        # GridFieldText(
        #   'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
        #   initially_hidden=True, field_name='location__owner__name', formatter='detail',
        #   extra='"role":"input/location"', editable=False
        #   ),
        # GridFieldText(
        #   'location__source', title=string_concat(_('location'), ' - ', _('source')),
        #   initially_hidden=True, editable=False
        #   ),
        # GridFieldLastModified(
        #   'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
        #   initially_hidden=True, editable=False
        #   ),
    )


class ItemDistributionList(GridReport):
    '''
    A list report to show item distribution.
    '''
    title = _("item distributions")
    basequeryset = ItemDistribution.objects.all()
    model = ItemDistribution
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/distribution/item-distributions.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/itemdistribution"'),

        # GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
        #               extra='"role":"input/item"', editable=False),

        # 新建一个显示列
        GridFieldText('item_display', title=_('distribute item'), field_name='item__nr', editable=False),

        # 因为是id 让外键永远不显示
        GridFieldText('item_id', title=_('item_id'), field_name='item_id', editable=False, hidden=True),

        GridFieldText('origin_display', title=_('distribute origin location'), field_name='origin__nr',
                      editable=False),
        GridFieldText('origin_id', title=_('origin_id'), field_name='origin_id', editable=False, hidden=True),

        GridFieldText('destination_display', title=_('distribute destination location'), field_name='destination__name',
                      editable=False),
        GridFieldText('destination_id', title=_('destination_id'), field_name='destination_id', editable=False,
                      hidden=True),

        GridFieldCurrency('cost', title=_('cost'), editable=False),

        GridFieldNumber('load_time', title=_('load time'), editable=False),
        GridFieldNumber('transit_time', title=_('transit time'), editable=False),
        GridFieldNumber('receive_time', title=_('receive time'), editable=False),

        GridFieldNumber('size_minimum', title=_('distribute size minimum'), editable=False),
        GridFieldNumber('size_multiple', title=_('distribute size multiple')),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        # GridFieldDuration('fence', title=_('fence'), initially_hidden=True),

        GridFieldText('resource_display', title=_('resource_display'), field_name='resource__nr', editable=False),
        GridFieldText('resource', title=_('resource_id'), field_name='resource_id', editable=False,
                      hidden=True),

        GridFieldNumber('resource_qty', title=_('resource quantity'), editable=False),

        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),

        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # GridFieldText('source', title=_('source'), initially_hidden=True),

        # GridFieldLastModified('lastmodified'),
        # Optional fields referencing the item
        # GridFieldText(
        #     'item__description', title=string_concat(_('item'), ' - ', _('description')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'item__category', title=string_concat(_('item'), ' - ', _('category')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
        #     field_name='item__owner__name', initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'item__source', title=string_concat(_('item'), ' - ', _('source')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldLastModified(
        #     'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
        #     initially_hidden=True, editable=False
        # ),
        # # Optional fields referencing the location
        # GridFieldText(
        #     'location__description', title=string_concat(_('location'), ' - ', _('description')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__category', title=string_concat(_('location'), ' - ', _('category')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'location__available', title=string_concat(_('location'), ' - ', _('available')),
        #     initially_hidden=True, field_name='location__available__name', formatter='detail',
        #     extra='"role":"input/calendar"', editable=False
        # ),
        # GridFieldText(
        #     'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
        #     initially_hidden=True, field_name='location__owner__name', formatter='detail',
        #     extra='"role":"input/location"', editable=False
        # ),
        # GridFieldText(
        #     'location__source', title=string_concat(_('location'), ' - ', _('source')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldLastModified(
        #     'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
        #     initially_hidden=True, editable=False
        # ),
        # # Optional fields referencing the origin location
        # GridFieldText(
        #     'origin__description', title=string_concat(_('origin'), ' - ', _('description')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'origin__category', title=string_concat(_('origin'), ' - ', _('category')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'origin__subcategory', title=string_concat(_('origin'), ' - ', _('subcategory')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldText(
        #     'origin__available', title=string_concat(_('origin'), ' - ', _('available')),
        #     initially_hidden=True, field_name='origin__available__name', formatter='detail',
        #     extra='"role":"input/calendar"', editable=False
        # ),
        # GridFieldText(
        #     'origin__owner', title=string_concat(_('origin'), ' - ', _('owner')),
        #     initially_hidden=True, field_name='origin__owner__name', formatter='detail',
        #     extra='"role":"input/location"', editable=False
        # ),
        # GridFieldText(
        #     'origin__source', title=string_concat(_('origin'), ' - ', _('source')),
        #     initially_hidden=True, editable=False
        # ),
        # GridFieldLastModified(
        #     'origin__lastmodified', title=string_concat(_('origin'), ' - ', _('last modified')),
        #     initially_hidden=True, editable=False
        # ),
    )


# CMARK 枚举API
class EnumView(View):
    def get(self, request, *args, **kwargs):
        if 'type' in kwargs and 'value' in kwargs:
            type = kwargs['type']
            value = kwargs['value']
            if type == 'item_status_by_type':
                t = Item.type_status[value]
                dic = la_enum.tuple2select(t) if t != None else None

                return HttpResponse(json.dumps(dic, cls=DjangoJSONEncoder),
                                    content_type='application/json')
        else:
            return HttpResponseBadRequest()


class ItemList(GridReport):
    '''
    A list report to show items.
    '''
    title = _("items")
    basequeryset = Item.objects.all()
    model = Item
    frozenColumns = 1
    editable = True
    help_url = 'user-guide/modeling-wizard/master-data/items.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/item"'),
        GridFieldText('nr', title=_('item nr'), editable=False),
        GridFieldText('name', title=_('name'), editable=False),
        GridFieldText('barcode', title=_('barcode'), editable=False),
        GridFieldChoice('type', title=_('type'), choices=Item.types, editable=False),
        GridFieldText('status', field_name='status', title=_('status'), editable=False),
        GridFieldChoice('plan_strategy', title=_('plan strategy'), choices=Item.strategies, editable=False),
        GridFieldChoice('lock_type', title=_('lock type'), choices=Item.lock_types, editable=False),
        GridFieldDate('lock_expire_at', title=_('lock expire at'), editable=False),
        GridFieldChoice('price_abc', title=_('price abc'), choices=Item.abc_types, editable=False),
        GridFieldChoice('qty_abc', title=_('qty abc'), choices=Item.abc_types, editable=False),
        GridFieldCurrency('cost', title=_('cost'), editable=False),
        GridFieldText('source', title=_('source'), editable=False, initially_hidden=True),
        # 新建一个显示列
        GridFieldText('owner_display', title=_('owner_display'), field_name='owner__nr', editable=False),
        # 因为是id 让外键永远不显示
        GridFieldText('owner', title=_('owner_id'), field_name='owner_id', editable=False, hidden=True),
        GridFieldNumber('gross_weight', title=_('gross weight'), editable=False),
        GridFieldNumber('net_weight', title=_('net weight'), editable=False),
        GridFieldText('physical_unit', title=_('physical unit'), editable=False),
        GridFieldText('project_nr', title=_('project nr'), editable=False, initially_hidden=True),
        GridFieldInteger('mpq', title=_('mpq'), editable=False),
        GridFieldInteger('outer_package_num', title=_('outer package num'), editable=False, initially_hidden=True),
        GridFieldInteger('pallet_num', title=_('pallet num'), editable=False, initially_hidden=True),
        GridFieldNumber('outer_package_gross_weight', title=_('outer package gross weight'), editable=False,
                        initially_hidden=True),
        GridFieldNumber('pallet_gross_weight', title=_('pallet gross weight'), editable=False, initially_hidden=True),
        GridFieldNumber('outer_package_volume', title=_('outer package volume'), editable=False, initially_hidden=True),
        GridFieldNumber('pallet_volume', title=_('pallet volume'), editable=False, initially_hidden=True),
        GridFieldDate('plan_list_date', title=_('plan list date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_delist_date', title=_('plan delist date'), editable=False, initially_hidden=True),
        GridFieldText('category', title=_('category'), editable=False, initially_hidden=True),
        GridFieldText('subcategory', title=_('subcategory'), editable=False, initially_hidden=True),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # CMARK 必须有为了弹框查询
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),
    )


# TODO 物料详情
class ItemDetail(View):
    def get(self, request, *args, **kwargs):
        # 默认为main
        # main:         主数据
        # supplier:     供应商
        # plan:         计划
        # simulation: 　模拟
        # forecast:     预测
        template_name = "input/item/detail_base.html"
        id = kwargs['id']
        return render(request, template_name, {'template_name': template_name, "date_types": Bucket.chioce_date_type()})


# 代号：GET_ITEM_MAIN_DATA_API
# 单个物料头部公共数据
class ItemMainData(View):
    def get(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = datetime.now()
        try:
            item = Item.objects.get(id=id)
        except:
            message.result = False
            message.code = 404
            message.message = "没有对应的物料"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')

        try:
            successor_nr = ItemSuccessor.objects.filter(item=item).order_by('priority').first().item_successor.nr
        except:
            successor_nr = None

        lock_types = {"current": item.lock_type, "values": la_enum.tuple2select(Item.lock_types)}

        item_statuses = {"current": item.status, "values": la_enum.tuple2select(Item.type_status[item.type])}

        plan_strategies = {"current": item.plan_strategy,
                           "values": la_enum.tuple2select(Item.strategies)}

        locations = Location.objects.select_related().all().order_by('id')
        location = []
        for f in locations:
            locationdict = {
                "id": f.id,
                "nr": f.nr,
                # TODO 暂时无数据
                "buffer": {
                    "total_qty": 0,
                    "available_qty": 0,
                    "buffer_price": 0
                }
            }
            location.append(locationdict)
        data = {
            "id": item.id,
            "nr": item.nr,
            "successor_nr": successor_nr,
            "description": item.description,
            "project_nr": item.project_nr,
            "location": location,
            "lock_types": lock_types,
            "lock_expire_at": item.lock_expire_at,
            "plan_strategies": plan_strategies,
            "statuses": item_statuses,
            "price_abc": item.price_abc,
            "qty_abc": item.qty_abc
        }
        message.result = True
        message.code = 200
        message.message = "相应数据查询成功"
        message.content = data
        return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                            content_type='application/json')

    def post(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        json_data = request.body
        data = json.loads(json_data)
        with transaction.atomic():
            # 创建保存点
            save_point = transaction.savepoint()
            try:
                item = Item.objects.get(id=id)
                item.description = data['description']
                item.project_nr = data['project_nr']
                if data['lock_types'] is None:
                    message.result = False
                    message.code = 200
                    message.message = "请选择锁定类型"
                    return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                        content_type='application/json')

                if data['lock_expire_at'] is None and data['lock_types'] == 'locked':
                    message.result = False
                    message.code = 200
                    message.message = "请选择到期时间"
                    return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                        content_type='application/json')

                item.lock_type = data['lock_types']
                item.lock_expire_at = data['lock_expire_at']
                item.save()

            except Exception as e:
                message.result = False
                message.code = 404
                message.message = "数据保存失败"
                transaction.savepoint_rollback(save_point)
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')
            else:
                transaction.savepoint_commit(save_point)
                message.result = True
                message.code = 200
                message.message = "数据保存成功"
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')


# 代号　GET_ITEM_SUPPLIERS_DATA_API
# 获取单个物料供应商界面主数据
class ItemSupplierData(View):
    def get(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        try:
            supplier = ItemSupplier.objects.all().order_by('priority', '-ratio', 'id').filter(item=id)
        except Exception as e:
            message.result = False
            message.code = 404
            message.message = "供应商不存在"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')
        data = []
        for f in supplier:
            supplier_dict = {
                "id": f.supplier.id,
                "name": f.supplier.name,
                "nr": f.supplier.nr,
                "country": f.supplier.country,
                "city": f.supplier.city,
                "address": f.supplier.address,
                "phone": f.supplier.phone,
                "telephone": f.supplier.telephone,
                "email": f.supplier.email,
                "contact": f.supplier.contact,
                "cost": decimal2float(f.cost),
                "cost_unit": decimal2float(f.cost_unit),
                "supplier_item_nr": f.supplier_item_nr
            }
            data.append(supplier_dict)
        message.result = True
        message.code = 200
        message.message = "相应数据查询成功"
        message.content = data
        return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                            content_type='application/json')


# 代号：GET_ITEM_MAIN_SUPPLIER_DATA_API
# 获取单个物料主数据页，前置期+供应商+包装部分
class MainSupplierData(View):
    def get(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = timezone.now()
        try:
            item = Item.objects.get(id=id)
        except Exception as e:
            message.result = False
            message.code = 404
            message.message = "没有对应的物料"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')
        try:
            item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                        effective_end__gte=current_time).order_by('priority', '-ratio',
                                                                                                  'id').first()
        except Exception as e:
            message.result = False
            message.code = 404
            message.message = "合法的主供应商不存在"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')

        lead_time = item_supplier.wd2cd()

        data = {
            "supplier_id": item_supplier.supplier.id,
            "name": item_supplier.supplier.name,
            "nr": item_supplier.supplier.nr,
            "product_time": decimal2float(item_supplier.product_time),
            "load_time": decimal2float(item_supplier.load_time),
            "transit_time": decimal2float(item_supplier.transit_time),
            "receive_time": decimal2float(item_supplier.receive_time),
            "plan_supplier_date": item_supplier.plan_supplier_date,
            "plan_load_date": item_supplier.plan_load_date,
            "plan_receive_date": item_supplier.plan_receive_date,
            "totall_lead_time": decimal2float(lead_time),
            "cost": decimal2float(item_supplier.cost),
            "cost_unit": decimal2float(item_supplier.cost_unit),
            "earliest_order_date": item_supplier.earliest_order_date,
            "lock_expire_at": item.lock_expire_at,
            "plan_list_date": item_supplier.plan_list_date,
            "plan_delist_date": item_supplier.plan_delist_date,
            "moq": decimal2float(item_supplier.moq),
            "mpq": decimal2float(item_supplier.mpq),
            "pallet_num": decimal2float(item_supplier.pallet_num),
            # TODO 手工MOQ暂时无数据
            "MOQ": 0,
            "order_unit_qty": decimal2float(item_supplier.order_unit_qty),
            "outer_package_num": decimal2float(item_supplier.outer_package_num),
            "order_max_qty": decimal2float(item_supplier.order_max_qty),
            "description": item_supplier.description
        }

        message.result = True
        message.code = 200
        message.message = "相应数据查询成功"
        message.content = data
        return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                            content_type='application/json')

    def post(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = datetime.now()
        json_data = request.body
        data = json.loads(json_data)
        with transaction.atomic():
            # 创建保存点
            save_point = transaction.savepoint()
            try:
                item = Item.objects.get(id=id)
                item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                            effective_end__gte=current_time).order_by('priority',
                                                                                                      '-ratio',
                                                                                                      'id').first()

                item_supplier.plan_supplier_date = data['plan_supplier_date']
                item_supplier.plan_load_date = data['plan_load_date']
                item_supplier.plan_receive_date = data['plan_receive_date']
                item_supplier.earliest_order_date = data['earliest_order_date']
                item.lock_expire_at = data['lock_expire_at']
                item_supplier.plan_list_date = data['plan_list_date']
                item_supplier.plan_delist_date = data['plan_delist_date']
                item_supplier.save()
                item.save()

            except Exception as e:
                message.result = False
                message.code = 404
                message.message = "数据保存失败"
                transaction.savepoint_rollback(save_point)
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')
            else:
                transaction.savepoint_commit(save_point)
                message.result = True
                message.code = 200
                message.message = "数据保存成功"
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')


# 代号：GET_ITEM_SIMULATION_DATA_API
# 获取单个物料模拟主数据列表
class ItemSimulation(View):
    def get(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = timezone.now()
        try:
            item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                   effective_end__gte=current_time).order_by('priority', '-ratio',
                                                                                             'id').first()
        except Exception as e:
            message.result = False
            message.code = 404
            message.message = "合法的主供应商不存在"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')

        data = {
            "supplier_id": item_supplier.supplier.id,
            "nr": item_supplier.supplier.nr,
            "name": item_supplier.supplier.name,
            # TODO 目前订货点没有数据
            "now_order_point": 0,
            "moq": decimal2float(item_supplier.moq),
            "order_max_qty": decimal2float(item_supplier.order_max_qty)
        }
        message.result = True
        message.code = 200
        message.message = "相应数据查询成功"
        message.content = data
        return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                            content_type='application/json')

    def post(self, request, id, *args, **kwargs):
        current_time = datetime.now()
        message = ResponseMessage()
        json_data = request.body
        data = json.loads(json_data)
        with transaction.atomic():
            # 创建保存点
            save_point = transaction.savepoint()
            try:
                item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                       effective_end__gte=current_time).order_by('priority', '-ratio',
                                                                                                 'id').first()

                item_supplier.mpq = data['mpq']
                item_supplier.outer_package_num = data['outer_package_num']
                item_supplier.pallet_num = data['pallet_num']
                item_supplier.moq = data['moq']
                item_supplier.save()

            except:
                message.result = False
                message.code = 404
                message.message = "数据保存失败"
                transaction.savepoint_rollback(save_point)
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')
            else:
                transaction.savepoint_commit(save_point)
                message.result = True
                message.code = 200
                message.message = "数据保存成功"
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')


# 代号：GET_ITEM_PLAN_DATA_API
# 获取单个物料计划主数据
class ItemPlan(View):
    def get(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = timezone.now()
        try:
            item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                   effective_end__gte=current_time).order_by('priority', '-ratio',
                                                                                             'id').first()
        except Exception as e:
            message.result = False
            message.code = 404
            message.message = "合法的主供应商不存在"
            return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                content_type='application/json')

        lead_time = item_supplier.wd2cd()

        data = {
            "supplier_id": item_supplier.supplier.id,
            "nr": item_supplier.supplier.nr,
            "name": item_supplier.supplier.name,
            "safe_buffer": 0,
            "moq": decimal2float(item_supplier.moq),
            "mpq": decimal2float(item_supplier.mpq),
            "outer_package_num": decimal2float(item_supplier.outer_package_num),
            "pallet_num": decimal2float(item_supplier.pallet_num),
            "lead_time": decimal2float(lead_time),
            "per_month_sale": 0,
            "last_year_sale": 0,
            "last_month_sale": 0,
            "now_month_sale": 0,
            "now_buffer_time": 0,
            "now_order_point": 0
        }

        message.result = True
        message.code = 200
        message.message = "相应数据查询成功"
        message.content = data
        return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                            content_type='application/json')

    def post(self, request, id, *args, **kwargs):
        message = ResponseMessage()
        current_time = datetime.now()
        json_data = request.body
        data = json.loads(json_data)
        safe_buffer = data['safe_buffer']
        with transaction.atomic():
            # 创建保存点
            save_point = transaction.savepoint()
            try:
                item_supplier = ItemSupplier.objects.filter(item=id, effective_start__lte=current_time,
                                                            effective_end__gte=current_time).order_by('priority', '-ratio',
                                                                                                 'id').first()
                item_supplier.mpq = data['mpq']
                item_supplier.outer_package_num = data['outer_package_num']
                item_supplier.pallet_num = data['pallet_num']
                item_supplier.moq = data['moq']
                item_supplier.save()

            except:
                message.result = False
                message.code = 404
                message.message = "数据保存失败"
                transaction.savepoint_rollback(save_point)
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')
            else:
                transaction.savepoint_commit(save_point)
                message.result = True
                message.code = 200
                message.message = "数据保存成功"
                return HttpResponse(json.dumps(message.__dict__, cls=DjangoJSONEncoder, ensure_ascii=False),
                                    content_type='application/json')


class ItemCustomerList(GridReport):
    '''
    A list report to show items.
    '''
    title = _("item customers")
    basequeryset = ItemCustomer.objects.all()
    model = ItemCustomer
    frozenColumns = 1
    editable = True

    rows = (
        # . Translators: Translation included with Django
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/itemcustomer"'),
        # 新建一个显示列
        GridFieldText('sale_item_display', title=_('sale_item_display'), field_name='sale_item__nr', editable=False),
        # 因为是id 让外键永远不显示
        GridFieldInteger('sale_item', title=_('sale_item_id'), field_name='sale_item_id', editable=False, hidden=True),

        GridFieldText('product_item_display', title=_('product_item_display'), field_name='product_item__nr',
                      editable=False),
        GridFieldInteger('product_item', title=_('product_item_id'), field_name='product_item_id', editable=False,
                         hidden=True),
        GridFieldText('customer_display', title=_('customer_display'), field_name='customer__nr', editable=False),
        GridFieldInteger('customer', title=_('customer_id'), field_name='customer_id', editable=False, hidden=True),
        GridFieldText('location_display', title=_('location_display'), field_name='location__nr', editable=False),
        GridFieldInteger('location', title=_('location_id'), field_name='location_id', editable=False, hidden=True),
        GridFieldText('customer_item_nr', title=_('customer item nr'), editable=False),
        GridFieldText('status', title=_('status'), editable=False),
        GridFieldChoice('lock_type', title=_('lock type'), choices=Item.lock_types, editable=False),
        GridFieldDate('lock_expire_at', title=_('lock expire at'), editable=False),
        GridFieldDate('plan_list_date', title=_('plan list date'), editable=False, initially_hidden=True),
        GridFieldDate('plan_delist_date', title=_('plan delist date'), editable=False, initially_hidden=True),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )


class ItemSuccessorList(GridReport):
    '''
    A list report to show items.
    '''
    title = _("item successors")
    basequeryset = ItemSuccessor.objects.all()
    model = ItemSuccessor
    frozenColumns = 1
    editable = True

    rows = (
        # . Translators: Translation included with Django
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/itemsuccessor"'),
        # 新建一个显示列
        GridFieldText('item_display', title=_('item_display'), field_name='item__nr', editable=False),
        # 因为是id 让外键永远不显示
        GridFieldText('item', title=_('item_id'), field_name='item_id', editable=False, hidden=True),
        GridFieldText('item_successor_display', title=_('item_successor_display'), field_name='item_successor__nr',
                      editable=False),
        # 因为是id 让外键永远不显示
        GridFieldText('item_successor', title=_('item_successor'), field_name='item_successor_id', editable=False,
                      hidden=True),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        GridFieldNumber('ratio', title=_('ratio'),
                        extra='"formatoptions":{"suffix":" %","defaultValue":"100.00"}', editable=False),

        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )


class SkillList(GridReport):
    '''
    A list report to show skills.
    '''
    title = _("skills")
    basequeryset = Skill.objects.all()
    model = Skill
    frozenColumns = 1
    help_url = 'user-guide/model-reference/skills.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/skill"',
                      editable=False),
        GridFieldText('nr', title=_('nr'), editable=False),

        GridFieldText('name', title=_('name'), editable=False),
        GridFieldText('category', title=_('category'), editable=False),
        GridFieldText('subcategory', title=_('subcategory'), editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),

    )


class ResourceSkillList(GridReport):
    '''
    A list report to show resource skills.
    '''
    title = _("resource skills")
    basequeryset = ResourceSkill.objects.all()
    model = ResourceSkill
    frozenColumns = 1
    help_url = 'user-guide/model-reference/resource-skills.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/resourceskill"', editable=False),
        GridFieldText('resource_display', title=_('resource_display'), field_name='resource__nr', formatter='detail',
                      extra='"role":"input/resource"', editable=False),
        GridFieldText('resource', title=_('resource_id'), field_name='resource_id', formatter='detail',
                      extra='"role":"input/resource"', hidden=True, editable=False),

        GridFieldText('skill_display', title=_('skill_display'), field_name='skill__nr', formatter='detail',
                      extra='"role":"input/skill"', editable=False),
        GridFieldText('skill', title=_('skill_id'), field_name='skill_id', formatter='detail',
                      extra='"role":"input/skill"', editable=False, hidden=True),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

    )


class OperationResourceList(GridReport):
    '''
    A list report to show operationresources.
    '''
    title = _("operation resources")
    basequeryset = OperationResource.objects.all()
    model = OperationResource
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/manufacturing-capacity/operation-resources.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/operationresource"', editable=False),
        GridFieldText('resource_display', title=_('resource_display'), field_name='resource__nr', editable=False),
        GridFieldText('resource', title=_('resource_id'), field_name='resource_id', hidden=True, editable=False),

        GridFieldText('operation_display', title=_('operation_display'), field_name='operation__nr', editable=False),
        GridFieldText('operation', title=_('operation_id'), field_name='operation_id', hidden=True, editable=False),

        GridFieldNumber('quantity', title=_('quantity'), editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),

        GridFieldText('skill_display', title=_('skill_display'), field_name='skill__nr', editable=False),
        GridFieldText('skill', title=_('skill_id'), field_name='skill_id', hidden=True, editable=False),

        GridFieldText('setup', title=_('setup'), editable=False),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldChoice('alternative_process_mode', title=_('alternative process mode'), choices=Operation.modes,
                        editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )

    # . Translators: Translation included with Django
    #     GridFieldText('name', title=_('name'), initially_hidden=True),
    #     GridFieldChoice('search', title=_('search mode'), choices=searchmode, initially_hidden=True),
    #     GridFieldText('source', title=_('source')),
    #     GridFieldLastModified('lastmodified'),
    #     # Operation fields
    #     GridFieldText('operation__description', title=string_concat(_('operation'), ' - ', _('description')),
    #                   initially_hidden=True, editable=False),
    #     GridFieldText('operation__category', title=string_concat(_('operation'), ' - ', _('category')),
    #                   initially_hidden=True, editable=False),
    #     GridFieldText('operation__subcategory', title=string_concat(_('operation'), ' - ', _('subcategory')),
    #                   initially_hidden=True, editable=False),
    #     GridFieldChoice('operation__type', title=string_concat(_('operation'), ' - ', _('type')),
    #                     choices=Operation.types, initially_hidden=True, editable=False),
    #     GridFieldDuration('operation__duration', title=string_concat(_('operation'), ' - ', _('duration')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldDuration('operation__duration_per', title=string_concat(_('operation'), ' - ', _('duration per unit')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldDuration('operation__fence', title=string_concat(_('operation'), ' - ', _('release fence')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldDuration('operation__posttime', title=string_concat(_('operation'), ' - ', _('post-op time')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldNumber('operation__sizeminimum', title=string_concat(_('operation'), ' - ', _('size minimum')),
    #                     initially_hidden=True, editable=False),
    #     GridFieldNumber('operation__sizemultiple', title=string_concat(_('operation'), ' - ', _('size multiple')),
    #                     initially_hidden=True, editable=False),
    #     GridFieldNumber('operation__sizemaximum', title=string_concat(_('operation'), ' - ', _('size maximum')),
    #                     initially_hidden=True, editable=False),
    #     GridFieldInteger('operation__priority', title=string_concat(_('operation'), ' - ', _('priority')),
    #                      initially_hidden=True, editable=False),
    #     GridFieldDateTime('operation__effective_start',
    #                       title=string_concat(_('operation'), ' - ', _('effective start')), initially_hidden=True,
    #                       editable=False),
    #     GridFieldDateTime('operation__effective_end', title=string_concat(_('operation'), ' - ', _('effective end')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldCurrency('operation__cost', title=string_concat(_('operation'), ' - ', _('cost')),
    #                       initially_hidden=True, editable=False),
    #     GridFieldChoice('operation__search', title=string_concat(_('operation'), ' - ', _('search mode')),
    #                     choices=searchmode, initially_hidden=True, editable=False),
    #     GridFieldText('operation__source', title=string_concat(_('operation'), ' - ', _('source')),
    #                   initially_hidden=True, editable=False),
    #     GridFieldLastModified('operation__lastmodified', title=string_concat(_('operation'), ' - ', _('last modified')),
    #                           initially_hidden=True, editable=False),
    # )


class OperationMaterialList(GridReport):
    '''
    A list report to show operationmaterials.
    '''
    title = _("operation materials")
    basequeryset = OperationMaterial.objects.all()
    model = OperationMaterial
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/manufacturing-bom/operation-materials.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/operationmaterial"', editable=False),

        GridFieldText('operation_display', title=_('operation_display'), field_name='operation__nr', editable=False),
        GridFieldText('item_display', title=_('item_display'), field_name='item__nr', editable=False),

        GridFieldInteger('operation', title=_('operation_id'), field_name='operation_id', formatter='detail',
                         editable=False, hidden=True),
        GridFieldInteger('item', title=_('item_id'), field_name='item_id', formatter='detail', editable=False,
                         hidden=True),

        GridFieldChoice('type', title=_('type'), choices=OperationMaterial.types, editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        GridFieldNumber('quantity', title=_('quantity'), editable=False),
        GridFieldNumber('quantity_fixed', title=_('fixed quantity'), editable=False),
        GridFieldNumber('materialbatch_per', title=_('materialbatch per'), editable=False),

        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldChoice('alternative_process_mode', title=_('alternative process mode'), choices=Operation.modes,
                        editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='id', editable=False, hidden=True),

        # . Translators: Translation included with Django
        # GridFieldText('name', title=_('name'), initially_hidden=True),
        # GridFieldChoice('search', title=_('search mode'), choices=searchmode, initially_hidden=True),
        # GridFieldText('source', title=_('source')),
        # GridFieldLastModified('lastmodified'),
        # GridFieldNumber('transferbatch', title=_('transfer batch quantity'), initially_hidden=True),
        # # Operation fields
        # GridFieldText('operation__description', title=string_concat(_('operation'), ' - ', _('description')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('operation__category', title=string_concat(_('operation'), ' - ', _('category')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('operation__subcategory', title=string_concat(_('operation'), ' - ', _('subcategory')),
        #               initially_hidden=True, editable=False),
        # GridFieldChoice('operation__type', title=string_concat(_('operation'), ' - ', _('type')),
        #                 choices=Operation.types, initially_hidden=True, editable=False),
        # GridFieldDuration('operation__duration', title=string_concat(_('operation'), ' - ', _('duration')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__duration_per', title=string_concat(_('operation'), ' - ', _('duration per unit')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__fence', title=string_concat(_('operation'), ' - ', _('release fence')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__posttime', title=string_concat(_('operation'), ' - ', _('post-op time')),
        #                   initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizeminimum', title=string_concat(_('operation'), ' - ', _('size minimum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizemultiple', title=string_concat(_('operation'), ' - ', _('size multiple')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizemaximum', title=string_concat(_('operation'), ' - ', _('size maximum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldInteger('operation__priority', title=string_concat(_('operation'), ' - ', _('priority')),
        #                  initially_hidden=True, editable=False),
        # GridFieldDateTime('operation__effective_start',
        #                   title=string_concat(_('operation'), ' - ', _('effective start')), initially_hidden=True,
        #                   editable=False),
        # GridFieldDateTime('operation__effective_end', title=string_concat(_('operation'), ' - ', _('effective end')),
        #                   initially_hidden=True, editable=False),
        # GridFieldCurrency('operation__cost', title=string_concat(_('operation'), ' - ', _('cost')),
        #                   initially_hidden=True, editable=False),
        # GridFieldChoice('operation__search', title=string_concat(_('operation'), ' - ', _('search mode')),
        #                 choices=searchmode, initially_hidden=True, editable=False),
        # GridFieldText('operation__source', title=string_concat(_('operation'), ' - ', _('source')),
        #               initially_hidden=True, editable=False),
        # GridFieldLastModified('operation__lastmodified', title=string_concat(_('operation'), ' - ', _('last modified')),
        #                       initially_hidden=True, editable=False),
    )


class ForecastYearList(GridReport):
    # template = ''
    title = _("forecastyears")
    basequeryset = ForecastYear.objects.all()
    model = ForecastYear
    frozenColumns = 1
    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/forecastyear"', editable=False),
        GridFieldText('item_display', title=_('item_display'), field_name='item__nr', editable=False),
        GridFieldText('item', title=_('item_id'), field_name='item_id', editable=False, hidden=True),
        GridFieldText('location_display', title=_('location'), field_name='location__nr', editable=False),
        GridFieldText('location', title=_('location_id'), field_name='location_id', editable=False, hidden=True),
        GridFieldText('customer_display', title=_('customer'), field_name='customer__nr', editable=False),
        GridFieldText('customer', title=_('customer_id'), field_name='customer_id', editable=False, hidden=True),
        GridFieldText('year', title=_('year'), editable=False),
        GridFieldInteger('date_number', title=_('date_number'), editable=False),
        GridFieldText('date_type', title=_('date_type'), editable=False),
        GridFieldNumber('ratio', title=_('forecast ratio'),
                        extra='"formatoptions":{"suffix":" %","defaultValue":"100.00"}', editable=False),
        GridFieldNumber('normal_qty', title=_('normal qty'), editable=False),
        GridFieldNumber('new_product_plan_qty', title=_('new product plan qty'), editable=False),
        GridFieldNumber('promotion_qty', title=_('promotion qty'), editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )


class ForecastVersionView(GridReport):
    title = _("forecastversions")
    basequeryset = ForecastVersion.objects.all().order_by('-created_at')
    model = ForecastVersion
    frozenColumns = 1
    template = 'input/forecastversion.html'

    # CMARK 设置默认的排序字段, 这个方法不是很好
    default_sort = None

    rows = (
        # GridFieldText('id', title=_('id'), editable=False),
        GridFieldText('nr', title=_('version nr'), key=True, formatter='customer',
                      extra='"role":"/data/input/forecast/?version_nr="', editable=False),
        GridFieldChoice('status', title=_('status'), choices=ForecastCommentOperation.statuses, editable=False),
        GridFieldText('status_value', title=_('status_value'), field_name='status', editable=False, hidden=True,
                      search=False),
        GridFieldText('create_user_display', title=_('create_user_display'), field_name='create_user__username',
                      editable=False),
        GridFieldText('create_user', title=_('create_user_id'), field_name='create_user_id', editable=False,
                      hidden=True, search=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='nr', editable=False, hidden=True, search=False),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True, search=False),

    )

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        data = {
            "date_types": ForecastYear.date_types
        }
        return data

    @method_decorator(staff_member_required)
    def post(self, request, *args, **kwargs):
        if request.FILES and len(request.FILES) == 1:
            excel_count = 0
            for filename, file in request.FILES.items():
                if file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    excel_count += 1
            if excel_count == 0:
                message = ResponseMessage(message='no excel file')
                return HttpResponse(json.dumps(message.__dict__, ensure_ascii=False), content_type='application/json')
            else:
                return HttpResponse(
                    json.dumps(ForecastUploader.upload_excel(request, Forecast).__dict__, ensure_ascii=False),
                    content_type='application/json')
        # 上传文件
        else:
            message = ResponseMessage(message='no excel file or file size>1')
            return HttpResponse(json.dumps(message.__dict__, ensure_ascii=False), content_type='application/json')

    @classmethod
    def _generate_spreadsheet_data(reportclass, request, output, *args, **kwargs):
        # 下载excel
        wb = Workbook()
        # 第一个sheet是ws,不然会自动生成一个sheet表
        ws = wb.worksheets[0]
        title = force_text(Forecast._meta.verbose_name or Forecast.title)
        ws.title = title
        headerstyle = NamedStyle(name="headerstyle")
        headerstyle.fill = PatternFill(fill_type="solid", fgColor='70c4f4')
        wb.add_named_style(headerstyle)
        nr = request.GET.get('nr', None)
        if nr:
            download_forecast = Forecast.objects.select_related('item', 'location', 'customer').filter(
                version_id=nr).order_by('-version_id', 'year', 'date_number')
        else:
            download_forecast = Forecast.objects.select_related('item', 'location', 'customer').all().order_by(
                '-version_id', 'year', 'date_number')

        if not download_forecast:
            return HttpResponse('没有下载数据')
        else:
            report_headers = (
                _('id'), _('item'), _('location'), _('customer'), _('year'), _('date_number'), _('date_type'),
                _('forecast ratio'),
                _('normal qty'), _('new product plan qty'), _('promotion qty'), _('status'), _('create user'),
                _('version'), _('created_at'), _('updated_at'))
            # 写入表头数据
            header = []
            for field in report_headers:
                cell = WriteOnlyCell(ws, value=force_text(field).title())
                cell.style = 'headerstyle'
                header.append(cell)
            ws.append(header)
            for f in download_forecast:
                body = [f.id, f.item.nr, f.location.nr, f.customer.nr if f.customer else None, f.year, f.date_number,
                        f.date_type, f.ratio,
                        f.normal_qty, f.new_product_plan_qty, f.promotion_qty, f.status, f.create_user.username,
                        f.version.nr, f.created_at, f.updated_at]
                ws.append(body)

            wb.save(output)

    @classmethod
    def _generate_csv_data(reportclass, request, *args, **kwargs):

        # 下载 csv
        sf = StringIO()
        nr = request.GET.get('nr', None)
        if nr:
            download_forecast = Forecast.objects.filter(version_id=nr).order_by('-version_id', 'year', 'date_number')
        else:
            download_forecast = Forecast.objects.select_related().all().order_by('-version_id', 'year', 'date_number')

        decimal_separator = get_format('DECIMAL_SEPARATOR', request.LANGUAGE_CODE, True)
        if decimal_separator == ",":
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=';')
        else:
            writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')

        reprot_headers = (
            _('id'), _('item'), _('location'), _('customer'), _('year'), _('date_number'), _('date_type'),
            _('forecast ratio'),
            _('normal qty'), _('new product plan qty'), _('promotion qty'), _('status'), _('create user'),
            _('version'), _('created_at'), _('updated_at'))

        writer.writerow(reprot_headers)

        for f in download_forecast:
            body = [f.id, f.item.nr, f.location.nr, f.customer.nr, f.year, f.date_number, f.date_type, f.ratio,
                    f.normal_qty, f.new_product_plan_qty, f.promotion_qty, f.status, f.create_user.username,
                    f.version.nr, f.created_at, f.updated_at]
            writer.writerow(body)
        return sf.getvalue()


class ForecastList(GridReport):
    # template = ''
    title = _("forecasts")
    # 版本倒排, 时间正排, 不考虑id
    basequeryset = Forecast.objects.all().order_by('-version_id', 'year', 'date_number')
    model = Forecast
    frozenColumns = 1
    template = 'input/forecast.html'

    default_sort = None

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        data = {
            "date_types": ForecastYear.date_types
        }
        return data

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, formatter='detail',
                         extra='"role":"input/forecast"', editable=False),
        GridFieldText('item_display', title=_('item_display'), field_name='item__nr', editable=False),
        GridFieldText('item', title=_('location_id'), field_name='location_id', editable=False, hidden=True,
                      search=False),
        GridFieldText('location_display', title=_('location_display'), field_name='location__nr', editable=False),
        GridFieldText('location', title=_('location_id'), field_name='location_id', editable=False, hidden=True,
                      search=False),
        GridFieldText('customer_display', title=_('customer_display'), field_name='customer__nr', editable=False),
        GridFieldText('customer', title=_('customer_id'), field_name='customer_id', editable=False, hidden=True,
                      search=False),
        GridFieldText('year', title=_('year'), editable=False),
        GridFieldInteger('date_number', title=_('date_number'), editable=False),
        GridFieldText('date_type', title=_('date_type'), editable=False),
        GridFieldNumber('ratio', title=_('forecast ratio'),
                        extra='"formatoptions":{"suffix":" %","defaultValue":"100.00"}', editable=False),
        GridFieldNumber('normal_qty', title=_('normal qty'), editable=False),
        GridFieldNumber('new_product_plan_qty', title=_('new product plan qty'), editable=False),
        GridFieldNumber('promotion_qty', title=_('promotion qty'), editable=False),
        GridFieldChoice('status', title=_('status'), choices=ForecastCommentOperation.statuses, editable=False),
        GridFieldText('status_value', title=_('status_value'), field_name='status', editable=False, hidden=True,
                      search=False),
        GridFieldText('create_user_display', title=_('create_user_display'), field_name='create_user__username',
                      editable=False),
        GridFieldText('create_user', title=_('create_user_id'), field_name='create_user_id', editable=False,
                      hidden=True),
        GridFieldText('version', title=_('version'), field_name='version', hidden=True, editable=False),
        GridFieldText('version_nr', title=_('version nr'), field_name='version__nr', editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )


# TODO 预测/预测版本的备注
class ForecastCommentView(View):
    # CMARK 免除csrf
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    @method_decorator(staff_member_required)
    def get(self, request, *args, **kwargs):
        # 根据Forecast, ForecastVersion 获取comment
        # request
        content_type_parameter = request.GET['content_type']
        content_id = request.GET['content_id']

        content_type = ContentType.objects.filter(app_label='input', model=request.GET['content_type'].lower()).first()

        if content_type:
            fields = [f.name for f in Comment._meta.fields]
            fields.append('user__username')
            comments = []
            for c in Comment.objects.filter(content_type=content_type, object_pk=content_id).order_by('-id').values(
                    *fields):
                # 翻译
                for f in Comment._meta.fields:
                    if f.choices is not None and len(f.choices) > 0:
                        c[f.name] = _(c[f.name])
                comments.append(c)
            return HttpResponse(json.dumps(comments,
                                           ensure_ascii=False,
                                           cls=DjangoJSONEncoder), content_type='application/json')
        else:
            return HttpResponseBadRequest('parameter is not correct')
        return HttpResponse(json.dumps([]))

    def operate(self, request, operation, content_type_parameter, content_type, content_id, comment):

        message = ResponseMessage(result=True)
        content_object = None
        if content_type_parameter == 'forecast':
            content_object = Forecast.objects.get(id=content_id)
        elif content_type_parameter == 'forecastversion':
            content_object = ForecastVersion.objects.get(nr=content_id)

        if content_type is None or content_object is None:
            message.result = False
            message.message = 'parameter error'
        else:
            if 'operation_forecast_ok' == operation:
                # 审批
                if content_object.can_ok():
                    content_object.status = 'ok'
                    if isinstance(content_object, ForecastVersion):
                        Forecast.objects.filter(version=content_object,
                                                status__in=ForecastCommentOperation.can_ok_status).update(status='ok')
                    content_object.save()
                else:
                    message.result = False
                    message.message = '状态不可进行审批操作'
            elif 'operation_forecast_nok' == operation:
                # 打回
                if content_object.can_nok():
                    content_object.status = 'nok'
                    if isinstance(content_object, ForecastVersion):
                        Forecast.objects.filter(version=content_object,
                                                status__in=ForecastCommentOperation.can_nok_status).update(status='nok')
                    content_object.save()
                else:
                    message.result = False
                    message.message = '状态不可进行打回操作'
            elif 'operation_forecast_cancel' == operation:
                if content_object.can_cancel():
                    content_object.status = 'cancel'
                    if isinstance(content_object, ForecastVersion):
                        Forecast.objects.filter(version=content_object,
                                                status__in=ForecastCommentOperation.can_cancel_status).update(
                            status='cancel')
                    content_object.save()
                else:
                    message.result = False
                    message.message = '状态不可进行取消操作'
            elif 'operation_forecast_release' == operation:
                if content_object.can_release():
                    content_object.status = 'release'
                    if isinstance(content_object, ForecastVersion):
                        Forecast.objects.filter(version=content_object,
                                                status__in=ForecastCommentOperation.can_release_status).update(
                            status='release')
                    content_object.save()
                else:
                    message.result = False
                    message.message = '状态不可进行审批操作'
            else:
                message.result = False
                message.message = 'operation参数错误, 不存在'

            if message.result:
                # 创建comment
                comment = Comment(user=request.user, content_type=content_type,
                                  content_object=content_object, comment=comment, operation=operation)
                comment.save()
                message.result = True

        return message

    @method_decorator(staff_member_required)
    def post(self, request, *args, **kwargs):

        try:
            data = json.JSONDecoder().decode(request.read().decode(request.encoding or settings.DEFAULT_CHARSET))

            content_type_parameter = data['content_type'].lower()
            content_type = ContentType.objects.filter(app_label='input',
                                                      model=content_type_parameter).first()
            operation = data['operation']
            message = ResponseMessage
            with transaction.atomic(using=request.database, savepoint=False):
                if 'content_id' in data:
                    content_id = data['content_id']
                    message = self.operate(request, operation, content_type_parameter, content_type, content_id,
                                           data['comment'])
                elif 'content_ids' in data:
                    for content_id in data['content_ids']:
                        message = self.operate(request, operation, content_type_parameter, content_type, content_id,
                                               data['comment'])
                    if len(data['content_ids']) > 1:
                        message.result = True
                        message.message = None
            return HttpResponse(json.dumps(message.__dict__, ensure_ascii=False), content_type='application/json')
        except ObjectDoesNotExist as e:
            print(e)
            traceback.print_exc()
            return HttpResponseBadRequest("parameter error, " + str(e), content_type='application/json')
        except Exception as e:
            print(e)
            traceback.print_exc()
            return HttpResponseServerError("server error, " + str(e), content_type='application/json')

    #
    # @classmethod
    # def _generate_spreadsheet_data(reportclass, request, output, *args, **kwargs):
    #     i=1


class DemandList(GridReport):
    '''
    A list report to show sales orders.
    '''
    template = 'input/demand.html'
    title = _("sales orders")
    basequeryset = Demand.objects.all()
    model = Demand
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/master-data/sales-orders.html'

    rows = (
        # . Translators: Translation included with Django

        GridFieldText('name', title=_('name'), key=True, formatter='detail', extra='"role":"input/demand"'),
        GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
                      extra='"role":"input/item"'),
        GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldText('customer', title=_('customer'), field_name='customer__name', formatter='detail',
                      extra='"role":"input/customer"'),
        GridFieldChoice('status', title=_('status'), choices=Demand.demandstatus),
        GridFieldNumber('quantity', title=_('quantity')),
        GridFieldDateTime('due', title=_('due')),
        GridFieldDuration('delay', title=_('delay'), editable=False, extra='"formatter":delayfmt'),
        GridFieldNumber('plannedquantity', title=_('planned quantity'), editable=False,
                        extra='"formatoptions":{"defaultValue":""}, "cellattr":plannedquantitycellattr'),
        GridFieldDateTime('deliverydate', title=_('delivery date'), editable=False),
        GridFieldText('description', title=_('description'), initially_hidden=True),
        GridFieldText('category', title=_('category'), initially_hidden=True),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True),
        GridFieldText('operation', title=_('delivery operation'), field_name='operation__name', formatter='detail',
                      extra='"role":"input/operation"', initially_hidden=True),
        GridFieldInteger('priority', title=_('priority')),
        GridFieldText('owner', title=_('owner'), field_name='owner__name', formatter='detail',
                      extra='"role":"input/demand"', initially_hidden=True),
        GridFieldDuration('maxlateness', title=_('maximum lateness'), initially_hidden=True),
        GridFieldNumber('minshipment', title=_('minimum shipment'), initially_hidden=True),
        GridFieldText('source', title=_('source')),
        GridFieldLastModified('lastmodified'),
        # Optional fields referencing the item
        GridFieldText(
            'item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldCurrency(
            'item__cost', title=string_concat(_('item'), ' - ', _('cost')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the location
        GridFieldText(
            'location__description', title=string_concat(_('location'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__category', title=string_concat(_('location'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__available', title=string_concat(_('location'), ' - ', _('available')),
            initially_hidden=True, field_name='location__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
            initially_hidden=True, field_name='location__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'location__source', title=string_concat(_('location'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the customer
        GridFieldText(
            'customer__description', title=string_concat(_('customer'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'customer__category', title=string_concat(_('customer'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'customer__subcategory', title=string_concat(_('customer'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'customer__owner', title=string_concat(_('customer'), ' - ', _('owner')),
            initially_hidden=True, field_name='customer__owner__name', formatter='detail',
            extra='"role":"input/customer"', editable=False
        ),
        GridFieldText(
            'customer__source', title=string_concat(_('customer'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'customer__lastmodified', title=string_concat(_('customer'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
    )

    if settings.ERP_CONNECTOR:
        actions = [{
            "name": 'erp_incr_export',
            "label": format_lazy("export to {erp}", erp=settings.ERP_CONNECTOR),
            "function": "ERPconnection.SODepExport(jQuery('#grid'),'SO')"
        }]
    else:
        actions = [
            {
                "name": 'inquiry',
                "label": format_lazy(_("change status to {status}"), status=_("inquiry")),
                "function": "grid.setStatus('inquiry')"
            },
            {
                "name": 'quote',
                "label": format_lazy(_("change status to {status}"), status=_("quote")),
                "function": "grid.setStatus('quote')"
            },
            {
                "name": 'open',
                "label": format_lazy(_("change status to {status}"), status=_("open")),
                "function": "grid.setStatus('open')"
            },
            {
                "name": 'closed',
                "label": format_lazy(_("change status to {status}"), status=_("closed")),
                "function": "grid.setStatus('closed')"},
            {
                "name": 'canceled',
                "label": format_lazy(_("change status to {status}"), status=_("canceled")),
                "function": "grid.setStatus('canceled')"
            },
        ]


class CalendarList(GridReport):
    '''
    A list report to show calendars.
    '''
    title = _("calendars")
    basequeryset = Calendar.objects.all()
    model = Calendar
    frozenColumns = 1
    help_url = 'user-guide/model-reference/calendars.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('name', title=_('name'), key=True, formatter='detail', extra='"role":"input/calendar"',
                      editable=False),
        GridFieldText('source', title=_('source'), editable=False),
        GridFieldNumber('defaultvalue', title=_('default value'), editable=False),
        GridFieldText('description', title=_('description'), editable=False),
        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),

        # CMARK 必须有为了弹框查询
        GridFieldText('_pk', field_name='name', editable=False, hidden=True),
        GridFieldText('_nk', field_name='name', editable=False, hidden=True),
    )


class CalendarBucketList(GridReport):
    '''
    A list report to show calendar buckets.
    '''
    title = _("calendar buckets")
    basequeryset = CalendarBucket.objects.all()
    model = CalendarBucket
    frozenColumns = 3
    help_url = 'user-guide/model-reference/calendars.html'

    rows = (
        GridFieldInteger('id', title=_('identifier'), formatter='detail', extra='"role":"input/calendarbucket"',
                         initially_hidden=True, editable=False),
        GridFieldText('calendar', title=_('calendar'), field_name='calendar__name', formatter='detail',
                      extra='"role":"input/calendar"', editable=False),
        GridFieldDateTime('startdate', title=_('start date'), editable=False),
        GridFieldDateTime('enddate', title=_('end date'), editable=False),
        GridFieldNumber('value', title=_('value'), editable=False),
        GridFieldInteger('priority', title=_('priority'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('monday', title=_('Monday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('tuesday', title=_('Tuesday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('wednesday', title=_('Wednesday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('thursday', title=_('Thursday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('friday', title=_('Friday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('saturday', title=_('Saturday'), editable=False),
        # . Translators: Translation included with Django
        GridFieldBool('sunday', title=_('Sunday'), editable=False),
        GridFieldDateTime('starttime', title=_('start time'), editable=False),
        GridFieldDateTime('endtime', title=_('end time'), editable=False),
        GridFieldText('source', title=_('source'), editable=False),
        # Not really right, since the engine doesn't read or store it
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
    )


class OperationList(GridReport):
    '''
    A list report to show operations.
    '''
    title = _("operations")
    basequeryset = Operation.objects.all()
    model = Operation
    frozenColumns = 1
    help_url = 'user-guide/modeling-wizard/manufacturing-bom/operations.html'

    rows = (
        # . Translators: Translation included with Django
        GridFieldText('id', title=_('id'), key=True, formatter='detail', extra='"role":"input/operation"',
                      editable=False),
        GridFieldText('nr', title=_('nr'), editable=False),
        GridFieldText('name', title=_('name'), editable=False),
        GridFieldChoice('type', title=_('type'), choices=Operation.types, editable=False),
        GridFieldText('location_display', title=_('location_display'), field_name='location__nr', editable=False),
        GridFieldText('location', title=_('location_id'), field_name='location_id', editable=False, hidden=True),
        GridFieldText('category', title=_('category'), initially_hidden=True, editable=False),
        GridFieldText('subcategory', title=_('subcategory'), initially_hidden=True, editable=False),
        GridFieldNumber('min_num_per', title=_('min num per'), editable=False),
        GridFieldNumber('max_num_per', title=_('max num per'), editable=False),
        GridFieldNumber('multiple_per', title=_('multiple per'), editable=False),
        GridFieldNumber('cost_per', title=_('cost per'), editable=False),
        GridFieldDuration('duration_per', title=_('duration per'), editable=False),
        GridFieldText('available', title=_('available'), field_name='available__name', formatter='detail',
                      extra='"role":"input/calendar"', editable=False),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),
        GridFieldChoice('alternative_process_mode', title=_('alternative process mode'), choices=Operation.modes,
                        editable=False),
        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        GridFieldText('_pk', field_name='id', editable=False, hidden=True),
        GridFieldText('_nk', field_name='nr', editable=False, hidden=True),

    )


class SubOperationList(GridReport):
    '''
    A list report to show suboperations.
    '''
    title = _("suboperations")
    basequeryset = SubOperation.objects.all()
    model = SubOperation
    frozenColumns = 1
    help_url = 'user-guide/model-reference/suboperations.html'

    rows = (
        GridFieldInteger('id', title=_('id'), key=True, editable=False, formatter='detail',
                         extra='"role":"input/suboperation"'),
        GridFieldText('parent_operation_display', title=_('parent_operation_display'),
                      field_name='parent_operation__nr', editable=False),
        GridFieldText('parent_operation', title=_('parent_operation_id'), field_name='parent_operation_id', hidden=True,
                      editable=False),

        GridFieldText('suboperation_display', title=_('suboperation_display'), field_name='suboperation__nr',
                      editable=False),
        GridFieldText('suboperation', title=_('suboperation_id'), field_name='suboperation_id', hidden=True,
                      editable=False),

        GridFieldInteger('priority', title=_('priority')),
        GridFieldDateTime('effective_start', title=_('effective start'), editable=False),
        GridFieldDateTime('effective_end', title=_('effective end'), editable=False),

        GridFieldCreateOrUpdateDate('created_at', title=_('created_at'), editable=False),
        GridFieldCreateOrUpdateDate('updated_at', title=_('updated_at'), editable=False),
        # GridFieldText('source', title=_('source')),
        # Operation fields
        # GridFieldText('operation__description', title=string_concat(_('operation'), ' - ', _('description')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('operation__category', title=string_concat(_('operation'), ' - ', _('category')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('operation__subcategory', title=string_concat(_('operation'), ' - ', _('subcategory')),
        #               initially_hidden=True, editable=False),
        # GridFieldChoice('operation__type', title=string_concat(_('operation'), ' - ', _('type')),
        #                 choices=Operation.types, initially_hidden=True, editable=False),
        # GridFieldDuration('operation__duration', title=string_concat(_('operation'), ' - ', _('duration')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__duration_per', title=string_concat(_('operation'), ' - ', _('duration per unit')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__fence', title=string_concat(_('operation'), ' - ', _('release fence')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('operation__posttime', title=string_concat(_('operation'), ' - ', _('post-op time')),
        #                   initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizeminimum', title=string_concat(_('operation'), ' - ', _('size minimum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizemultiple', title=string_concat(_('operation'), ' - ', _('size multiple')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('operation__sizemaximum', title=string_concat(_('operation'), ' - ', _('size maximum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldInteger('operation__priority', title=string_concat(_('operation'), ' - ', _('priority')),
        #                  initially_hidden=True, editable=False),
        # GridFieldDateTime('operation__effective_start',
        #                   title=string_concat(_('operation'), ' - ', _('effective start')), initially_hidden=True,
        #                   editable=False),
        # GridFieldDateTime('operation__effective_end', title=string_concat(_('operation'), ' - ', _('effective end')),
        #                   initially_hidden=True, editable=False),
        # GridFieldCurrency('operation__cost', title=string_concat(_('operation'), ' - ', _('cost')),
        #                   initially_hidden=True, editable=False),
        # GridFieldChoice('operation__search', title=string_concat(_('operation'), ' - ', _('search mode')),
        #                 choices=searchmode, initially_hidden=True, editable=False),
        # GridFieldText('operation__source', title=string_concat(_('operation'), ' - ', _('source')),
        #               initially_hidden=True, editable=False),
        # GridFieldLastModified('operation__lastmodified', title=string_concat(_('operation'), ' - ', _('last modified')),
        #                       initially_hidden=True, editable=False),
        # # Suboperation fields
        # GridFieldText('suboperation__description', title=string_concat(_('suboperation'), ' - ', _('description')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('suboperation__category', title=string_concat(_('suboperation'), ' - ', _('category')),
        #               initially_hidden=True, editable=False),
        # GridFieldText('suboperation__subcategory', title=string_concat(_('suboperation'), ' - ', _('subcategory')),
        #               initially_hidden=True, editable=False),
        # GridFieldChoice('suboperation__type', title=string_concat(_('suboperation'), ' - ', _('type')),
        #                 choices=Operation.types, initially_hidden=True, editable=False),
        # GridFieldDuration('suboperation__duration', title=string_concat(_('suboperation'), ' - ', _('duration')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('suboperation__duration_per',
        #                   title=string_concat(_('suboperation'), ' - ', _('duration per unit')), initially_hidden=True,
        #                   editable=False),
        # GridFieldDuration('suboperation__fence', title=string_concat(_('suboperation'), ' - ', _('release fence')),
        #                   initially_hidden=True, editable=False),
        # GridFieldDuration('suboperation__posttime', title=string_concat(_('suboperation'), ' - ', _('post-op time')),
        #                   initially_hidden=True, editable=False),
        # GridFieldNumber('suboperation__sizeminimum', title=string_concat(_('suboperation'), ' - ', _('size minimum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('suboperation__sizemultiple', title=string_concat(_('suboperation'), ' - ', _('size multiple')),
        #                 initially_hidden=True, editable=False),
        # GridFieldNumber('suboperation__sizemaximum', title=string_concat(_('suboperation'), ' - ', _('size maximum')),
        #                 initially_hidden=True, editable=False),
        # GridFieldInteger('suboperation__priority', title=string_concat(_('suboperation'), ' - ', _('priority')),
        #                  initially_hidden=True, editable=False),
        # GridFieldDateTime('suboperation__effective_start',
        #                   title=string_concat(_('suboperation'), ' - ', _('effective start')), initially_hidden=True,
        #                   editable=False),
        # GridFieldDateTime('suboperation__effective_end',
        #                   title=string_concat(_('suboperation'), ' - ', _('effective end')), initially_hidden=True,
        #                   editable=False),
        # GridFieldCurrency('suboperation__cost', title=string_concat(_('suboperation'), ' - ', _('cost')),
        #                   initially_hidden=True, editable=False),
        # GridFieldChoice('suboperation__search', title=string_concat(_('suboperation'), ' - ', _('search mode')),
        #                 choices=searchmode, initially_hidden=True, editable=False),
        # GridFieldText('suboperation__source', title=string_concat(_('suboperation'), ' - ', _('source')),
        #               initially_hidden=True, editable=False),
        # GridFieldLastModified('suboperation__lastmodified',
        #                       title=string_concat(_('suboperation'), ' - ', _('last modified')), initially_hidden=True,
        #                       editable=False),
    )


class OperationPlanMixin:
    if 'freppledb.inventoryplanning' in settings.INSTALLED_APPS:
        segmentlist = Segment.segmentList

    @classmethod
    def operationplanExtraBasequery(cls, query, request):
        if 'freppledb.inventoryplanning' in settings.INSTALLED_APPS:
            segmentname = request.prefs.get('segment', None) if request.prefs else None
            if segmentname:
                try:
                    segment = Segment.objects.all().using(request.database).get(pk=segmentname)
                    query = query.extra(
                        where=[
                            "exists ( %s and operationplan.item_id = item.name and operationplan.destination_id = location.nr)" % segment.getQuery()]
                    )
                except Segment.DoesNotExist:
                    pass
        if 'freppledb.forecast' in settings.INSTALLED_APPS:
            return query.extra(select={
                'demand': '''
          select json_agg(json_build_array(value, key, tp))
          from (
            select
              key, value,
              case when demand.name is not null then 'D' when forecast.name is not null then 'F' end as tp
            from jsonb_each_text(operationplan.plan->'pegging')
            left outer join demand on key = demand.name
            left outer join forecast on substring(key from 0 for position(' - ' in key)) = forecast.name
            where demand.name is not null or forecast.name is not null
            order by value desc, key desc
            limit 10
          ) peg''',
                'end_items': '''
          select json_agg(json_build_array(key, val))
          from (
            select coalesce(demand.item_id, forecast.item_id) as key, sum(value::numeric) as val
            from jsonb_each_text(operationplan.plan->'pegging')
            left outer join demand on key = demand.name
            left outer join forecast on substring(key from 0 for position(' - ' in key)) = forecast.name
            group by coalesce(demand.item_id, forecast.item_id)
            order by 2 desc
            limit 10
            ) peg_items'''
            })
        else:
            return query.extra(select={
                'demand': '''
          select json_agg(json_build_array(value, key))
          from (
            select key, value
            from jsonb_each_text(operationplan.plan->'pegging')
            order by value desc, key desc
            limit 10
            ) peg''',
                'end_items': '''
          select json_agg(json_build_array(key, val))
          from (
            select demand.item_id as key, sum(value::numeric) as val
            from jsonb_each_text(operationplan.plan->'pegging')
            inner join demand on key = demand.name
            group by demand.item_id
            order by 2 desc
            limit 10
            ) peg_items'''
            })


class ManufacturingOrderList(OperationPlanMixin, GridReport):
    '''
    A list report to show manufacturing orders.
    '''
    template = 'input/operationplanreport.html'
    title = _("manufacturing orders")
    default_sort = (2, 'desc')
    model = ManufacturingOrder
    frozenColumns = 2
    multiselect = True
    editable = True
    height = 250
    help_url = 'user-guide/modeling-wizard/manufacturing-bom/manufacturing-orders.html'

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        if args and args[0]:
            request.session['lasttab'] = 'manufacturingorders'
            return {
                'active_tab': 'manufacturingorders',
                'model': Location,
                'title': force_text(Location._meta.verbose_name) + " " + args[0],
                'post_title': _('manufacturing orders')
            }
        else:
            return {'active_tab': 'manufacturingorders'}

    @classmethod
    def basequeryset(reportclass, request, *args, **kwargs):
        q = ManufacturingOrder.objects.all()
        if args and args[0]:
            q = q.filter(location=args[0])
        q = reportclass.operationplanExtraBasequery(q, request)
        return q.extra(select={
            'material': "(select json_agg(json_build_array(item_id, quantity)) from (select item_id, round(quantity,2) quantity from operationplanmaterial where operationplan_id = operationplan.id order by quantity limit 10) mat)",
            'resource': "(select json_agg(json_build_array(resource_id, quantity)) from (select resource_id, round(quantity,2) quantity from operationplanresource where operationplan_id = operationplan.id order by quantity desc limit 10) res)",
            'setup_duration': "(operationplan.plan->'setup')",
            'setup_end': "(operationplan.plan->>'setupend')",
            'feasible': "coalesce((operationplan.plan->>'feasible')::boolean, true)",
        })

    rows = (
        GridFieldInteger('id', title=_('identifier'), key=True, formatter='detail',
                         extra="role:'input/manufacturingorder'", initially_hidden=True),
        GridFieldText('reference', title=_('reference'), editable=not settings.ERP_CONNECTOR),
        GridFieldNumber('color', title=_('inventory status'), formatter='color', width='125', editable=False,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldText('operation__item__name', title=_('item'), formatter='detail', extra='"role":"input/item"'),
        GridFieldText('operation__location__name', title=_('location'), formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldText('operation', title=_('operation'), field_name='operation__name', formatter='detail',
                      extra='"role":"input/operation"'),
        GridFieldDateTime('startdate', title=_('start date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"min"'),
        GridFieldDateTime('enddate', title=_('end date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"max"'),
        GridFieldNumber('quantity', title=_('quantity'),
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"sum"'),
        GridFieldChoice('status', title=_('status'), choices=OperationPlan.orderstatus,
                        editable=not settings.ERP_CONNECTOR),
        GridFieldNumber('criticality', title=_('criticality'), editable=False, initially_hidden=True,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldDuration('delay', title=_('delay'), editable=False, initially_hidden=True,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'),
        GridFieldText('demand', title=_('demands'), editable=False, search=False, sortable=False,
                      formatter='demanddetail', extra='"role":"input/demand"'),
        GridFieldText('material', title=_('materials'), editable=False, search=False, sortable=False,
                      initially_hidden=True, formatter='listdetail', extra='"role":"input/item"'),
        GridFieldText('resource', title=_('resources'), editable=False, search=False, sortable=False,
                      initially_hidden=True, formatter='listdetail', extra='"role":"input/resource"'),
        GridFieldInteger('owner', title=_('owner'), f2018103011304720181030113047ield_name='owner__id',
                         extra='"formatoptions":{"defaultValue":""}',
                         initially_hidden=True),
        GridFieldText('source', title=_('source')),
        GridFieldLastModified('lastmodified'),
        GridFieldText('operation__description', title=string_concat(_('operation'), ' - ', _('description')),
                      initially_hidden=True),
        GridFieldText('operation__category', title=string_concat(_('operation'), ' - ', _('category')),
                      initially_hidden=True),
        GridFieldText('operation__subcategory', title=string_concat(_('operation'), ' - ', _('subcategory')),
                      initially_hidden=True),
        GridFieldChoice('operation__type', title=string_concat(_('operation'), ' - ', _('type')),
                        choices=Operation.types, initially_hidden=True),
        GridFieldDuration('operation__duration', title=string_concat(_('operation'), ' - ', _('duration')),
                          initially_hidden=True),
        GridFieldDuration('operation__duration_per', title=string_concat(_('operation'), ' - ', _('duration per unit')),
                          initially_hidden=True),
        GridFieldDuration('operation__fence', title=string_concat(_('operation'), ' - ', _('release fence')),
                          initially_hidden=True),
        GridFieldDuration('operation__posttime', title=string_concat(_('operation'), ' - ', _('post-op time')),
                          initially_hidden=True),
        GridFieldNumber('operation__sizeminimum', title=string_concat(_('operation'), ' - ', _('size minimum')),
                        initially_hidden=True),
        GridFieldNumber('operation__sizemultiple', title=string_concat(_('operation'), ' - ', _('size multiple')),
                        initially_hidden=True),
        GridFieldNumber('operation__sizemaximum', title=string_concat(_('operation'), ' - ', _('size maximum')),
                        initially_hidden=True),
        GridFieldInteger('operation__priority', title=string_concat(_('operation'), ' - ', _('priority')),
                         initially_hidden=True),
        GridFieldDateTime('operation__effective_start',
                          title=string_concat(_('operation'), ' - ', _('effective start')), initially_hidden=True),
        GridFieldDateTime('operation__effective_end', title=string_concat(_('operation'), ' - ', _('effective end')),
                          initially_hidden=True),
        GridFieldCurrency('operation__cost', title=string_concat(_('operation'), ' - ', _('cost')),
                          initially_hidden=True),
        GridFieldChoice('operation__search', title=string_concat(_('operation'), ' - ', _('search mode')),
                        choices=searchmode, initially_hidden=True),
        GridFieldText('operation__source', title=string_concat(_('operation'), ' - ', _('source')),
                      initially_hidden=True),
        GridFieldLastModified('operation__lastmodified', title=string_concat(_('operation'), ' - ', _('last modified')),
                              initially_hidden=True),
        GridFieldDuration('setup_duration', title=_('setup time'), initially_hidden=True, search=False),
        GridFieldDateTime('setup_end', title=_('setup end date'), initially_hidden=True, search=False),
        GridFieldBool('feasible', title=_('feasible'), editable=False, initially_hidden=True, search=False),
        # Optional fields referencing the item
        GridFieldText(
            'operation__item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='operation__item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'operation__item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the location
        GridFieldText(
            'operation__location__description', title=string_concat(_('location'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__location__category', title=string_concat(_('location'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'operation__location__available', title=string_concat(_('location'), ' - ', _('available')),
            initially_hidden=True, field_name='operation__location__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'operation__location__owner', title=string_concat(_('location'), ' - ', _('owner')),
            initially_hidden=True, field_name='operation__location__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'operation__location__source', title=string_concat(_('location'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'operation__location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'end_items', title=_('end items'), editable=False, search=False, sortable=False,
            initially_hidden=True, formatter='listdetail', extra='"role":"input/item"'
        ),
    )

    if settings.ERP_CONNECTOR:
        actions = [{
            "name": 'erp_incr_export',
            "label": format_lazy("export to {erp}", erp=settings.ERP_CONNECTOR),
            "function": "ERPconnection.IncrementalExport(jQuery('#grid'),'MO')"
        }]
    else:
        actions = [
            {
                "name": 'proposed',
                "label": format_lazy(_("change status to {status}"), status=_("proposed")),
                "function": "grid.setStatus('proposed')"
            },
            {
                "name": 'approved',
                "label": format_lazy(_("change status to {status}"), status=_("approved")),
                "function": "grid.setStatus('approved')"
            },
            {
                "name": 'confirmed',
                "label": format_lazy(_("change status to {status}"), status=_("confirmed")),
                "function": "grid.setStatus('confirmed')"
            },
            {
                "name": 'closed',
                "label": format_lazy(_("change status to {status}"), status=_("closed")),
                "function": "grid.setStatus('closed')"
            },
        ]

    @classmethod
    def initialize(reportclass, request):
        if reportclass._attributes_added != 2:
            reportclass._attributes_added = 2
            # Adding custom operation attributes
            for f in getAttributeFields(Operation, related_name_prefix="operation"):
                f.editable = False
                reportclass.rows += (f,)
            # Adding custom location attributes
            for f in getAttributeFields(Location, related_name_prefix="location"):
                f.editable = False
                reportclass.rows += (f,)


class DistributionOrderList(OperationPlanMixin, GridReport):
    '''
    A list report to show distribution orders.
    '''
    template = 'input/operationplanreport.html'
    title = _("distribution orders")
    default_sort = (2, 'desc')
    basequeryset = DistributionOrder.objects.all()
    model = DistributionOrder
    frozenColumns = 2
    multiselect = True
    editable = True
    height = 250
    help_url = 'user-guide/modeling-wizard/distribution/distribution-orders.html'

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        if args and args[0]:
            path = request.path.split('/')[-2]
            if path == 'in':
                return {
                    'active_tab': 'inboundorders',
                    'model': Location,
                    'title': force_text(DistributionOrder._meta.verbose_name) + " " + args[0],
                    'post_title': _('inbound distribution')
                }
            elif path == 'out':
                return {
                    'active_tab': 'outboundorders',
                    'model': Location,
                    'title': force_text(DistributionOrder._meta.verbose_name) + " " + args[0],
                    'post_title': _('outbound distribution')
                }
            else:
                return {'active_tab': 'edit'}
        else:
            return {'active_tab': 'edit'}

    @classmethod
    def basequeryset(reportclass, request, *args, **kwargs):
        q = DistributionOrder.objects.all()
        if args and args[0]:
            path = request.path.split('/')[-2]
            if path == 'out':
                q = q.filter(origin_id=args[0])
            elif path == 'in':
                q = q.filter(destination_id=args[0])
            else:
                q = q.filter(location=args[0])
        q = reportclass.operationplanExtraBasequery(q, request)
        return q.extra(select={
            'total_cost': "cost*quantity",
            'feasible': "coalesce((operationplan.plan->>'feasible')::boolean, true)",
        })

    rows = (
        GridFieldInteger('id', title=_('identifier'), key=True, formatter='detail',
                         extra='role:"input/distributionorder"'),
        GridFieldText('reference', title=_('reference'), editable=not settings.ERP_CONNECTOR),
        GridFieldNumber('color', title=_('inventory status'), formatter='color', width='125', editable=False,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
                      extra='"role":"input/item"'),
        GridFieldText('origin', title=_('origin'), field_name='origin__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldText('destination', title=_('destination'), field_name='destination__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldDateTime('startdate', title=_('shipping date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"min"'),
        GridFieldDateTime('enddate', title=_('receipt date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"max"'),
        GridFieldNumber('quantity', title=_('quantity'),
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"sum"'),
        GridFieldChoice('status', title=_('status'), choices=DistributionOrder.orderstatus,
                        editable=not settings.ERP_CONNECTOR),
        GridFieldCurrency(
            'item__cost', title=string_concat(_('item'), ' - ', _('cost')),
            editable=False, extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'
        ),
        GridFieldCurrency('total_cost', title=_('total cost'), editable=False, search=False,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"sum"'),
        GridFieldNumber('criticality', title=_('criticality'), editable=False, initially_hidden=True,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldDuration('delay', title=_('delay'), editable=False, initially_hidden=True,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'),
        GridFieldText('demand', title=_('demands'), editable=False, search=False, sortable=False,
                      formatter='demanddetail', extra='"role":"input/demand"'),
        GridFieldText('source', title=_('source')),
        GridFieldLastModified('lastmodified'),
        GridFieldBool('feasible', title=_('feasible'), editable=False, initially_hidden=True, search=False),
        # Optional fields referencing the item
        GridFieldText(
            'item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the origin location
        GridFieldText(
            'origin__description', title=string_concat(_('origin'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'origin__category', title=string_concat(_('origin'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'origin__subcategory', title=string_concat(_('origin'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'origin__available', title=string_concat(_('origin'), ' - ', _('available')),
            initially_hidden=True, field_name='origin__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'origin__owner', title=string_concat(_('origin'), ' - ', _('owner')),
            initially_hidden=True, field_name='origin__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'origin__source', title=string_concat(_('origin'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'origin__lastmodified', title=string_concat(_('origin'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the destination location
        GridFieldText(
            'destination__description', title=string_concat(_('destination'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'destination__category', title=string_concat(_('destination'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'destination__subcategory', title=string_concat(_('destination'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'destination__available', title=string_concat(_('destination'), ' - ', _('available')),
            initially_hidden=True, field_name='origin__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'destination__owner', title=string_concat(_('destination'), ' - ', _('owner')),
            initially_hidden=True, field_name='origin__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'destination__source', title=string_concat(_('destination'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'destination__lastmodified', title=string_concat(_('destination'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'end_items', title=_('end items'), editable=False, search=False, sortable=False,
            initially_hidden=True, formatter='listdetail', extra='"role":"input/item"'
        ),
    )

    if settings.ERP_CONNECTOR:
        actions = [{
            "name": 'erp_incr_export',
            "label": format_lazy("export to {erp}", erp=settings.ERP_CONNECTOR),
            "function": "ERPconnection.IncrementalExport(jQuery('#grid'),'DO')"
        }]

    else:
        actions = [
            {
                "name": 'proposed',
                "label": format_lazy(_("change status to {status}"), status=_("proposed")),
                "function": "grid.setStatus('proposed')"
            },
            {
                "name": 'approved',
                "label": format_lazy(_("change status to {status}"), status=_("approved")),
                "function": "grid.setStatus('approved')"
            },
            {
                "name": 'confirmed',
                "label": format_lazy(_("change status to {status}"), status=_("confirmed")),
                "function": "grid.setStatus('confirmed')"
            },
            {
                "name": 'closed',
                "label": format_lazy(_("change status to {status}"), status=_("closed")),
                "function": "grid.setStatus('closed')"
            },
        ]

    @classmethod
    def initialize(reportclass, request):
        if reportclass._attributes_added != 2:
            reportclass._attributes_added = 2
            # Adding custom item attributes
            for f in getAttributeFields(Item, related_name_prefix="item"):
                f.editable = False
                reportclass.rows += (f,)
            # Adding custom location attributes
            for f in getAttributeFields(Location, related_name_prefix="origin"):
                f.editable = False
                reportclass.rows += (f,)
            # Adding custom location attributes
            for f in getAttributeFields(Location, related_name_prefix="destination"):
                f.editable = False
                reportclass.rows += (f,)


class PurchaseOrderList(OperationPlanMixin, GridReport):
    '''
    A list report to show purchase orders.
    '''
    template = 'input/operationplanreport.html'
    title = _("purchase orders")
    model = PurchaseOrder
    default_sort = (2, 'desc')
    frozenColumns = 2
    multiselect = True
    editable = True
    height = 250
    help_url = 'user-guide/modeling-wizard/purchasing/purchase-orders.html'

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        if args and args[0]:
            request.session['lasttab'] = 'purchaseorders'
            path = request.path.split('/')[-3]
            if path == 'supplier':
                return {
                    'active_tab': 'purchaseorders',
                    'model': Supplier,
                    'title': force_text(Supplier._meta.verbose_name) + " " + args[0],
                    'post_title': _('purchase orders')
                }
            elif path == 'location':
                return {
                    'active_tab': 'purchaseorders',
                    'model': Location,
                    'title': force_text(Location._meta.verbose_name) + " " + args[0],
                    'post_title': _('purchase orders')
                }
            elif path == 'item':
                return {
                    'active_tab': 'purchaseorders',
                    'model': Item,
                    'title': force_text(Item._meta.verbose_name) + " " + args[0],
                    'post_title': _('purchase orders')
                }
        else:
            return {'active_tab': 'purchaseorders'}

    @classmethod
    def basequeryset(reportclass, request, *args, **kwargs):
        q = PurchaseOrder.objects.all()
        if args and args[0]:
            path = request.path.split('/')[-3]
            if path == 'supplier':
                try:
                    sup = Supplier.objects.all().using(request.database).get(name=args[0])
                    lft = sup.lft
                    rght = sup.rght
                except Supplier.DoesNotExist:
                    lft = 1
                    rght = 1
                q = q.filter(supplier__lft__gte=lft, supplier__rght__lte=rght)
            elif path == 'location':
                try:
                    loc = Location.objects.all().using(request.database).get(name=args[0])
                    lft = loc.lft
                    rght = loc.rght
                except Location.DoesNotExist:
                    lft = 1
                    rght = 1
                q = q.filter(location__lft__gte=lft, location__rght__lte=rght)
            elif path == 'item':
                try:
                    itm = Item.objects.all().using(request.database).get(name=args[0])
                    lft = itm.lft
                    rght = itm.rght
                except Item.DoesNotExist:
                    lft = 1
                    rght = 1
                q = q.filter(item__lft__gte=lft, item__rght__lte=rght)
        q = reportclass.operationplanExtraBasequery(q, request)
        return q.extra(select={
            'total_cost': "cost*quantity",
            'unit_cost': "coalesce((select max(cost) from itemsupplier where itemsupplier.item_id = operationplan.item_id and itemsupplier.location_id = operationplan.location_id and itemsupplier.supplier_id = operationplan.supplier_id), (select cost from item where item.name = operationplan.item_id))",
            'feasible': "coalesce((operationplan.plan->>'feasible')::boolean, true)",
        })

    rows = (
        GridFieldInteger('id', title=_('identifier'), key=True, formatter='detail', extra='role:"input/purchaseorder"'),
        GridFieldText('reference', title=_('reference'), editable=not settings.ERP_CONNECTOR),
        GridFieldNumber('color', title=_('inventory status'), formatter='color', width='125', editable=False,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
                      extra='"role":"input/item"'),
        GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldText('supplier', title=_('supplier'), field_name='supplier__name', formatter='detail',
                      extra='"role":"input/supplier"'),
        GridFieldDateTime('startdate', title=_('ordering date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"min"'),
        GridFieldDateTime('enddate', title=_('receipt date'),
                          extra='"formatoptions":{"srcformat":"Y-m-d H:i:s","newformat":"Y-m-d H:i:s", "defaultValue":""}, "summaryType":"max"'),
        GridFieldNumber('quantity', title=_('quantity'),
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"sum"'),
        GridFieldChoice('status', title=_('status'), choices=PurchaseOrder.orderstatus,
                        editable=not settings.ERP_CONNECTOR),
        GridFieldCurrency('unit_cost', title=string_concat(_('item'), ' - ', _('cost')), editable=False, search=False,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'),
        GridFieldCurrency('total_cost', title=_('total cost'), editable=False, search=False,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"sum"'),
        GridFieldNumber('criticality', title=_('criticality'), editable=False, initially_hidden=True,
                        extra='"formatoptions":{"defaultValue":""}, "summaryType":"min"'),
        GridFieldDuration('delay', title=_('delay'), editable=False, initially_hidden=True,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'),
        GridFieldText('demand', title=_('demands'), editable=False, search=False, sortable=False,
                      formatter='demanddetail', extra='"role":"input/demand"'),
        GridFieldText('source', title=_('source')),
        GridFieldBool('feasible', title=_('feasible'), editable=False, initially_hidden=True, search=False),
        GridFieldLastModified('lastmodified'),
        # Optional fields referencing the item
        GridFieldText(
            'item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the location
        GridFieldText(
            'location__description', title=string_concat(_('location'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__category', title=string_concat(_('location'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__available', title=string_concat(_('location'), ' - ', _('available')),
            initially_hidden=True, field_name='location__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
            initially_hidden=True, field_name='location__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'location__source', title=string_concat(_('location'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the supplier
        GridFieldText(
            'supplier__description', title=string_concat(_('supplier'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'supplier__category', title=string_concat(_('supplier'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'supplier__subcategory', title=string_concat(_('supplier'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'supplier__owner', title=string_concat(_('supplier'), ' - ', _('owner')),
            initially_hidden=True, field_name='supplier__owner__name', formatter='detail',
            extra='"role":"input/supplier"', editable=False
        ),
        GridFieldText(
            'supplier__source', title=string_concat(_('supplier'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'supplier__lastmodified', title=string_concat(_('supplier'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'end_items', title=_('end items'), editable=False, search=False, sortable=False,
            initially_hidden=True, formatter='listdetail', extra='"role":"input/item"'
        ),
    )

    if settings.ERP_CONNECTOR:
        actions = [{
            "name": 'erp_incr_export',
            "label": format_lazy("export to {erp}", erp=settings.ERP_CONNECTOR),
            "function": "ERPconnection.IncrementalExport(jQuery('#grid'),'PO')"
        }]
    else:
        actions = [
            {
                "name": 'proposed',
                "label": format_lazy(_("change status to {status}"), status=_("proposed")),
                "function": "grid.setStatus('proposed')"
            },
            {
                "name": 'approved',
                "label": format_lazy(_("change status to {status}"), status=_("approved")),
                "function": "grid.setStatus('approved')"
            },
            {
                "name": 'confirmed',
                "label": format_lazy(_("change status to {status}"), status=_("confirmed")),
                "function": "grid.setStatus('confirmed')"
            },
            {
                "name": 'closed',
                "label": format_lazy(_("change status to {status}"), status=_("closed")),
                "function": "grid.setStatus('closed')"
            },
        ]

    @classmethod
    def initialize(reportclass, request):
        if reportclass._attributes_added != 2:
            reportclass._attributes_added = 2
            # Adding custom item attributes
            for f in getAttributeFields(Item, related_name_prefix="item"):
                f.editable = False
                reportclass.rows += (f,)
            # Adding custom location attributes
            for f in getAttributeFields(Location, related_name_prefix="location"):
                f.editable = False
                reportclass.rows += (f,)
            # Adding custom supplier attributes
            for f in getAttributeFields(Supplier, related_name_prefix="supplier"):
                f.editable = False
                reportclass.rows += (f,)


class DeliveryOrderList(GridReport):
    '''
    A list report to show delivery plans for demand.
    '''
    template = 'input/deliveryorder.html'
    title = _("Delivery orders")
    model = DeliveryOrder
    frozenColumns = 0
    editable = True
    multiselect = True
    help_url = 'user-guide/model-reference/delivery-orders.html'
    rows = (
        # . Translators: Translation included with Django
        GridFieldInteger('id', title=_('identifier'), initially_hidden=True, key=True, formatter='detail',
                         extra='role:"input/deliveryorder"'),
        GridFieldText('reference', title=_('reference'), editable=not settings.ERP_CONNECTOR),
        GridFieldText('demand', title=_('demand'), field_name="demand__name", formatter='detail',
                      extra='"role":"input/demand"'),
        GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail',
                      extra='"role":"input/item"'),
        GridFieldText('customer', title=_('customer'), field_name='demand__customer__name', formatter='detail',
                      extra='"role":"input/customer"'),
        GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail',
                      extra='"role":"input/location"'),
        GridFieldNumber('quantity', title=_('quantity')),
        GridFieldNumber('demand__quantity', title=_('demand quantity'), editable=False),
        GridFieldDateTime('startdate', title=_('start date')),
        GridFieldDateTime('enddate', title=_('end date'),
                          extra=GridFieldDateTime.extra + ',"cellattr":enddatecellattr'),
        GridFieldDateTime('due', field_name='due', title=_('due date'), editable=False),
        GridFieldChoice('status', title=_('status'), choices=OperationPlan.orderstatus,
                        editable=not settings.ERP_CONNECTOR),
        GridFieldDuration('delay', title=_('delay'), editable=False, initially_hidden=True,
                          extra='"formatoptions":{"defaultValue":""}, "summaryType":"max"'),
        # Optional fields referencing the item
        GridFieldText(
            'item__description', title=string_concat(_('item'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__category', title=string_concat(_('item'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__subcategory', title=string_concat(_('item'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__owner', title=string_concat(_('item'), ' - ', _('owner')),
            field_name='item__owner__name', initially_hidden=True, editable=False
        ),
        GridFieldText(
            'item__source', title=string_concat(_('item'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'item__lastmodified', title=string_concat(_('item'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the location
        GridFieldText(
            'location__description', title=string_concat(_('location'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__category', title=string_concat(_('location'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__subcategory', title=string_concat(_('location'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'location__available', title=string_concat(_('location'), ' - ', _('available')),
            initially_hidden=True, field_name='location__available__name', formatter='detail',
            extra='"role":"input/calendar"', editable=False
        ),
        GridFieldText(
            'location__owner', title=string_concat(_('location'), ' - ', _('owner')),
            initially_hidden=True, field_name='location__owner__name', formatter='detail',
            extra='"role":"input/location"', editable=False
        ),
        GridFieldText(
            'location__source', title=string_concat(_('location'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'location__lastmodified', title=string_concat(_('location'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        ),
        # Optional fields referencing the customer
        GridFieldText(
            'demand__customer__description', title=string_concat(_('customer'), ' - ', _('description')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'demand__customer__category', title=string_concat(_('customer'), ' - ', _('category')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'demand__customer__subcategory', title=string_concat(_('customer'), ' - ', _('subcategory')),
            initially_hidden=True, editable=False
        ),
        GridFieldText(
            'demand__customer__owner', title=string_concat(_('customer'), ' - ', _('owner')),
            initially_hidden=True, field_name='supplier__owner__name', formatter='detail',
            extra='"role":"input/supplier"', editable=False
        ),
        GridFieldText(
            'demand__customer__source', title=string_concat(_('customer'), ' - ', _('source')),
            initially_hidden=True, editable=False
        ),
        GridFieldLastModified(
            'demand__customer__lastmodified', title=string_concat(_('customer'), ' - ', _('last modified')),
            initially_hidden=True, editable=False
        )
    )

    @classmethod
    def basequeryset(reportclass, request, *args, **kwargs):
        if args and args[0]:
            try:
                itm = Item.objects.all().using(request.database).get(name=args[0])
                lft = itm.lft
                rght = itm.rght
            except Item.DoesNotExist:
                lft = 1
                rght = 1
            return DeliveryOrder.objects.all().filter(item__lft__gte=lft, item__rght__lte=rght)
        else:
            return DeliveryOrder.objects.all()

    @classmethod
    def extra_context(reportclass, request, *args, **kwargs):
        if args and args[0]:
            request.session['lasttab'] = 'plandetail'
            return {
                'active_tab': 'plandetail',
                'title': force_text(Item._meta.verbose_name) + " " + args[0],
                'post_title': _("Delivery orders")
            }
        else:
            return {'active_tab': 'plandetail'}

    @classmethod
    def initialize(reportclass, request):
        if reportclass._attributes_added != 2:
            reportclass._attributes_added = 2
            # Adding custom item attributes
            for f in getAttributeFields(Item, related_name_prefix="item"):
                f.editable = False
                f.initially_hidden = True
                reportclass.rows += (f,)
            # Adding custom location attributes
            for f in getAttributeFields(Location, related_name_prefix="location"):
                f.editable = False
                f.initially_hidden = True
                reportclass.rows += (f,)
            # Adding custom customer attributes
            for f in getAttributeFields(Customer, related_name_prefix="demand__customer"):
                f.editable = False
                f.initially_hidden = True
                reportclass.rows += (f,)


class OperationPlanDetail(View):

    def getData(self, request):
        # Current date
        try:
            current_date = datetime.strptime(
                Parameter.objects.using(request.database).get(name="currentdate").value,
                "%Y-%m-%d %H:%M:%S"
            )
        except:
            current_date = datetime.now()
        cursor = connections[request.database].cursor()

        # Read the results from the database
        ids = request.GET.getlist('id')
        first = True
        if not ids:
            yield "[]"
            raise StopIteration
        try:
            opplans = [x for x in OperationPlan.objects.all().using(request.database).filter(id__in=ids).select_related(
                "operation")]
            opplanmats = [x for x in OperationPlanMaterial.objects.all().using(request.database).filter(
                operationplan__id__in=ids).values()]
            opplanrscs = [x for x in OperationPlanResource.objects.all().using(request.database).filter(
                operationplan__id__in=ids).values()]
        except Exception as e:
            logger.error("Error retrieving operationplan data: %s" % e)
            yield "[]"
            raise StopIteration

        # Store my permissions
        view_PO = request.user.has_perm("input.view_purchaseorder")
        view_MO = request.user.has_perm("input.view_manufacturingorder")
        view_DO = request.user.has_perm("input.view_distributionorder")
        view_OpplanMaterial = request.user.has_perm("input.view_operationplanmaterial")
        view_OpplanResource = request.user.has_perm("input.view_operationplanresource")

        # Loop over all operationplans
        for opplan in opplans:

            # Check permissions
            if opplan.type == "DO" and not view_DO:
                continue
            if opplan.type == "PO" and not view_PO:
                continue
            if opplan.type == "MO" and not view_MO:
                continue
            try:
                # Base information
                res = {
                    "id": opplan.id,
                    "start": opplan.startdate.strftime("%Y-%m-%dT%H:%M:%S") if opplan.startdate else None,
                    "end": opplan.enddate.strftime("%Y-%m-%dT%H:%M:%S") if opplan.enddate else None,
                    "quantity": float(opplan.quantity),
                    "criticality": float(opplan.criticality) if opplan.criticality else '',
                    "delay": opplan.delay.total_seconds() if opplan.delay else '',
                    "status": opplan.status,
                    "reference": opplan.reference,
                    "type": opplan.type,
                    "name": opplan.name,
                    "destination": opplan.destination_id,
                    "location": opplan.location_id,
                    "origin": opplan.origin_id,
                    "supplier": opplan.supplier_id,
                    "item": opplan.item_id,
                    "color": float(opplan.color) if opplan.color else ''
                }
                if opplan.plan and 'pegging' in opplan.plan:
                    res["pegging_demand"] = []
                    for d, q in opplan.plan['pegging'].items():
                        try:
                            obj = Demand.objects.all().using(request.database).only("name", "item", "due").get(name=d)
                            dmd = obj.name
                            due = obj.due.strftime("%Y-%m-%dT%H:%M:%S")
                            item = obj.item.name
                        except Demand.DoesNotExist:
                            # Looks like this demand was deleted since the plan was generated
                            continue
                        res["pegging_demand"].append({
                            "demand": {"name": dmd, "item": {"name": item}, "due": due},
                            "quantity": q
                        })
                    res["pegging_demand"].sort(key=lambda f: (f['demand']['name'], f['demand']['due']))
                if opplan.operation:
                    res['operation'] = {
                        "name": opplan.operation.name,
                        "type": "operation_%s" % opplan.operation.type
                    }

                # Information on materials
                if view_OpplanMaterial:
                    firstmat = True
                    for m in opplanmats:
                        if m['operationplan_id'] != opplan.id:
                            continue
                        if firstmat:
                            firstmat = False
                            res['flowplans'] = []
                        res['flowplans'].append({
                            "date": m['flowdate'].strftime("%Y-%m-%dT%H:%M:%S"),
                            "quantity": float(m['quantity']),
                            "onhand": float(m['onhand'] or 0),
                            "buffer": {
                                "name": "%s @ %s" % (m['item_id'], m['location_id'])
                            }
                        })

                # Information on resources
                if view_OpplanResource:
                    firstres = True
                    for m in opplanrscs:
                        if m['operationplan_id'] != opplan.id:
                            continue
                        if firstres:
                            firstres = False
                            res['loadplans'] = []
                        res['loadplans'].append({
                            "date": m['startdate'].strftime("%Y-%m-%dT%H:%M:%S"),
                            "quantity": float(m['quantity']),
                            "resource": {
                                "name": m['resource_id']
                            }
                        })

                # Retrieve network status
                if opplan.item_id:
                    cursor.execute('''
            with items as (
               select name from item where name = %s
               )
            select
              items.name, false, location.nr, onhand.qty, orders_plus.PO,
              coalesce(orders_plus.DO, 0) - coalesce(orders_minus.DO, 0),
              orders_plus.MO, sales.BO, sales.SO
            from items
            cross join location
            left outer join (
              select item_id, location_id, onhand as qty
              from buffer
              inner join items on items.name = buffer.item_id
              ) onhand
            on onhand.item_id = items.name and onhand.location_id = location.nr
            left outer join (
              select item_id, coalesce(location_id, destination_id) as location_id,
              sum(case when type = 'MO' then quantity end) as MO,
              sum(case when type = 'PO' then quantity end) as PO,
              sum(case when type = 'DO' then quantity end) as DO
              from operationplan
              inner join items on items.name = operationplan.item_id
              and status in ('approved', 'confirmed')
              group by item_id, coalesce(location_id, destination_id)
              ) orders_plus
            on orders_plus.item_id = items.name and orders_plus.location_id = location.nr
            left outer join (
              select item_id, origin_id as location_id,
              sum(quantity) as DO
              from operationplan
              inner join items on items.name = operationplan.item_id
              and status in ('approved', 'confirmed')
              and type = 'DO'
              group by item_id, origin_id
              ) orders_minus
            on orders_minus.item_id = items.name and orders_minus.location_id = location.nr
            left outer join (
              select item_id, location_id,
              sum(case when due < %s then quantity end) as BO,
              sum(case when due >= %s then quantity end) as SO
              from demand
              inner join items on items.name = demand.item_id
              where status in ('open', 'quote')
              group by item_id, location_id
              ) sales
            on sales.item_id = items.name and sales.location_id = location.nr
            where
              onhand.qty is not null
              or orders_plus.MO is not null
              or orders_plus.PO is not null
              or orders_plus.DO is not null
              or orders_minus.DO is not null
              or sales.BO is not null
              or sales.SO is not null
            order by items.name, location.nr
            ''', (opplan.item_id, current_date, current_date))
                    res['network'] = []
                    for a in cursor.fetchall():
                        res['network'].append([
                            a[0], a[1], a[2],
                            float(a[3] or 0), float(a[4] or 0), float(a[5] or 0),
                            float(a[6] or 0), float(a[7] or 0), float(a[8] or 0)
                        ])

                # Final result
                if first:
                    yield "[%s" % json.dumps(res)
                    first = False
                else:
                    yield ',%s' % json.dumps(res)
                yield "]"
            except Exception as e:
                # Ignore exceptions and move on
                logger.error("Error retrieving operationplan: %s" % e)

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    @method_decorator(staff_member_required)
    def get(self, request):
        # Only accept ajax requests on this URL
        if not request.is_ajax():
            raise Http404('Only ajax requests allowed')

        # Stream back the response
        response = StreamingHttpResponse(
            content_type='application/json; charset=%s' % settings.DEFAULT_CHARSET,
            streaming_content=self.getData(request)
        )
        response['Cache-Control'] = "no-cache, no-store"
        return response

    @method_decorator(staff_member_required)
    def post(self, request):
        # Only accept ajax requests on this URL
        if not request.is_ajax():
            raise Http404('Only ajax requests allowed')

        # Parse the posted data
        try:
            data = json.JSONDecoder().decode(request.read().decode(request.encoding or settings.DEFAULT_CHARSET))
        except Exception as e:
            logger.error("Error updating operationplan data: %s" % e)
            return HttpResponseServerError("Error updating operationplan data", content_type='text/html')

        update_PO = request.user.has_perm("input.change_purchaseorder")
        update_MO = request.user.has_perm("input.change_manufacturingorder")
        update_DO = request.user.has_perm("input.change_distributionorder")

        for opplan_data in data:
            try:
                # Read the object from the database
                opplan = OperationPlan.objects.all().using(request.database).get(id=opplan_data.get('id', None))

                # Check permissions
                if opplan.type == "DO" and not update_DO:
                    continue
                if opplan.type == "PO" and not update_PO:
                    continue
                if opplan.type == "MO" and not update_MO:
                    continue

                # Update fields
                save = False
                if "start" in opplan_data:
                    # Update start date
                    opplan.startdate = datetime.strptime(opplan_data['start'], "%Y-%m-%dT%H:%M:%S")
                    save = True
                if "end" in opplan_data:
                    # Update end date
                    opplan.enddate = datetime.strptime(opplan_data['end'], "%Y-%m-%dT%H:%M:%S")
                    save = True
                if "quantity" in opplan_data:
                    # Update quantity
                    opplan.quantity = opplan_data['quantity']
                    save = True
                if "status" in opplan_data:
                    # Status quantity
                    opplan.status = opplan_data['status']
                    save = True

                if "reference" in opplan_data:
                    # Update reference
                    opplan.reference = opplan_data['reference']
                    save = True

                # Save if changed
                if save:
                    opplan.save(
                        using=request.database,
                        update_fields=["startdate", "enddate", "quantity", "reference", "lastmodified"]
                    )
            except OperationPlan.DoesNotExist:
                # Silently ignore
                pass
            except Exception as e:
                # Swallow the exception and move on
                logger.error("Error updating operationplan: %s" % e)
        return HttpResponse(content="OK")
