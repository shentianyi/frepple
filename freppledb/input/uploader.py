import logging
import traceback
from datetime import datetime

from django.db import transaction
from django.utils import timezone
from django.utils.encoding import force_text
from freppledb.common.models import User
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.writer.write_only import WriteOnlyCell
from io import BytesIO
from freppledb.common.dataload import parseExcelWorksheet
from freppledb.common.message.responsemessage import ResponseMessage
from freppledb.input.models import Forecast, Location, Item, Customer, ForecastVersion, ForecastCommentOperation
from openpyxl.cell import cell

logger = logging.getLogger(__name__)


class ForecastUploader:
    @classmethod
    def upload_excel(cls, request, model):
        message = ResponseMessage()

        try:
            for name, file in request.FILES.items():
                # print(name)
                # 读excel
                wb = load_workbook(filename=file, read_only=True, data_only=True)
                # 第一个sheet
                sheet = wb[wb.sheetnames[0]]
                row_count = 0
                row_number = 1
                # 对应表头是在第几列
                headers_index = {}
                # 对应表头与模型field的name
                headers_field_name = {}
                # 所有excel中的forecast
                forecasts = []

                starttime = datetime.now()

                locations = {}
                items = {}
                customers = {}

                for row in sheet.iter_rows():
                    row_count += 1
                    if row_number == 1:
                        row_number += 1
                        # required_fields = set()
                        # # 获取必须值 & 表头
                        # for field in model._meta.fields:
                        #     if not field.blank and field.default==NOT_PROVIDED and not isinstance(field, AutoField):
                        #      required_fields.add(field.name)

                        # 获取excelheader
                        _headers = [i.value for i in row]
                        index = 0
                        for h in _headers:
                            for field in model._meta.fields:
                                if h == field.name.lower or h == field.verbose_name.lower():
                                    headers_index[h] = index
                                    headers_field_name[h] = field.name
                            index += 1
                    # 取值
                    else:
                        values = [i.value for i in row]
                        none_len = 0
                        for v in values:
                            if v is None:
                                none_len += 1

                        if none_len == len(values):
                            continue

                        forecast = Forecast()
                        forecast.date_type = request.POST.get('date_type', 'W')

                        for k, v in headers_index.items():
                            value = values[v]
                            field_name = headers_field_name[k]

                            if field_name == 'location':
                                try:
                                    if value not in locations:
                                        locations[value] = \
                                        Location.objects.using(request.database).values('id').get(nr=value)['id']
                                    forecast.location_id = locations.get(value, None)
                                except Location.DoesNotExist as e:
                                    message.result = False
                                    message.message = "地点数据不存在"
                                    return message
                            elif field_name == 'item':
                                try:
                                    if value not in items:
                                        items[value] = Item.objects.using(request.database).values('id').get(nr=value)[
                                            'id']
                                    forecast.item_id = items[value]
                                except Item.DoesNotExist as e:
                                    message.result = False
                                    message.message = "物料数据不存在"
                                    return message

                            elif field_name == 'customer':
                                try:
                                    if value not in customers:
                                        customers[value] = \
                                        Customer.objects.using(request.database).values('id').get(nr=value)['id']
                                    forecast.customer_id = customers[value]
                                except Customer.DoesNotExist as e:
                                    message.result = False
                                    message.message = "客户数据不存在"
                                    return message
                            elif field_name == 'year':
                                forecast.year = value
                            elif field_name == 'date_number':
                                forecast.date_number = value
                            elif field_name == 'normal_qty':
                                forecast.normal_qty = value
                            elif field_name == 'ratio' and value is not None:
                                forecast.ratio = value
                            elif field_name == 'new_product_plan_qty':
                                forecast.new_product_plan_qty = value
                            elif field_name == 'promotion_qty':
                                forecast.promotion_qty = value

                        forecast.create_user = request.user
                        forecasts.append(forecast)

                endtime = datetime.now()

                logger.debug(
                    "start time vs end time: %s vs %s, %s" % (starttime, endtime, (endtime - starttime).microseconds))
                print("start time vs end time: %s vs %s, %s" % (starttime, endtime, (endtime - starttime).microseconds))

                if row_count < 2:
                    message.result = False
                    message.message = '文件没有数据'
                else:
                    with transaction.atomic(using=request.database, savepoint=False):
                        if request.POST['action'] == 'create':
                            # 创建 version
                            forecast_version = ForecastVersion()
                            forecast_version.create_user = request.user
                            forecast_version.created_at = timezone.now
                            # forecast_version.status = ForecastCommentOperation.statuses[0][0]
                            forecast_version.nr = timezone.now().strftime('%y%m%d%H%M%S')
                            forecast_version.save()
                            for f in forecasts:
                                f.version = forecast_version
                            # 创建forecast
                            Forecast.objects.bulk_create(forecasts, batch_size=100)
                        elif request.POST['action'] == 'update':
                            # 查找最新的version
                            forecast_version = ForecastVersion.objects.using(request.database).latest('created_at')
                            if forecast_version is None:
                                message.result = False
                                message.message = '版本不存在,不可以更新!'
                            else:
                                # excel 中传的字段
                                excel_fields = [v for k, v in headers_field_name.items()]

                                # 可以被更新的字段
                                update_fields = ['date_type', 'ratio', 'normal_qty', 'normal_qty',
                                                 'new_product_plan_qty', 'promotion_qty']
                                starttime = datetime.now()

                                for f in forecasts:
                                    # 更新
                                    try:
                                        update_forecast = Forecast.objects.using(request.database).filter(
                                            # version=forecast_version,
                                            version=forecast_version,
                                            location_id=f.location_id,
                                            item_id=f.item_id,
                                            customer_id=f.customer_id,
                                            year=f.year,
                                            date_type=f.date_type,
                                            date_number=f.date_number).latest('id')

                                        # 判断是否在excel中传了值
                                        for field in update_fields:
                                            if field in excel_fields:
                                                setattr(update_forecast, field, getattr(f, field, None))

                                        update_forecast.save()
                                    except Forecast.DoesNotExist as e:
                                        f.version = forecast_version
                                        f.save()

                                endtime = datetime.now()
                                print("start time vs end time: %s vs %s, %s" % (
                                    starttime, endtime, (endtime - starttime).microseconds))

                        message.result = True
                        message.message = '上传成功'
        except Exception as e:
            print(e)
            traceback.print_exc()
            message.result = False
            message.message = str(e)
        return message
